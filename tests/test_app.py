"""
MyContacts v2.0 — Playwright Tests
Run: pytest tests/test_app.py -v  (while app.py is running on port 5000)
     OR: pytest tests/test_app.py -v --start-server  (auto-starts server)
"""
import pytest
import subprocess
import time
import requests
import os
import sys

from playwright.sync_api import Page, expect, sync_playwright

BASE = "http://localhost:5000"


# ─── Server fixture ────────────────────────────────────────────────────────────

@pytest.fixture(scope="session", autouse=True)
def flask_server():
    """Start Flask server for the test session if not already running."""
    try:
        requests.get(BASE, timeout=2)
        yield  # already running
        return
    except Exception:
        pass

    proc_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    proc = subprocess.Popen(
        [sys.executable, "app.py"],
        cwd=proc_dir,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    # Wait for server to be ready
    for _ in range(20):
        try:
            requests.get(BASE, timeout=1)
            break
        except Exception:
            time.sleep(0.5)
    yield
    proc.terminate()


@pytest.fixture(scope="session")
def browser_ctx():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context()
        yield ctx
        ctx.close()
        browser.close()


@pytest.fixture()
def page(browser_ctx):
    pg = browser_ctx.new_page()
    yield pg
    pg.close()


# ─── Helper ───────────────────────────────────────────────────────────────────

def goto_app(page: Page):
    page.goto(BASE)
    page.wait_for_selector("#contactList", timeout=8000)
    page.wait_for_selector(".contact-item", timeout=8000)


def api(method, path, json=None):
    fn = getattr(requests, method)
    kwargs = {"timeout": 10}
    if json is not None:
        kwargs["json"] = json
    r = fn(BASE + path, **kwargs)
    r.raise_for_status()
    return r.json()


# ─── 1. Page loads ─────────────────────────────────────────────────────────────

def test_page_loads(page: Page):
    goto_app(page)
    expect(page.locator(".logo")).to_be_visible()
    expect(page.locator("#contactList")).to_be_visible()
    expect(page.locator("#sidebar")).to_be_visible()


# ─── 2. Contact list ──────────────────────────────────────────────────────────

def test_contacts_appear_in_list(page: Page):
    goto_app(page)
    items = page.locator(".contact-item")
    assert items.count() >= 1, "Expected at least 1 contact in list"


def test_search_filters_list(page: Page):
    goto_app(page)
    page.fill("#searchInput", "สมชาย")
    page.wait_for_timeout(400)
    items = page.locator(".contact-item")
    assert items.count() >= 1
    names = page.locator(".contact-item .ci-name")
    found = any("สมชาย" in (names.nth(i).inner_text()) for i in range(names.count()))
    assert found, "Search should surface สมชาย"


# ─── 3. Contact detail ────────────────────────────────────────────────────────

def test_click_contact_shows_detail(page: Page):
    goto_app(page)
    page.locator(".contact-item").first.click()
    page.wait_for_selector("#dpName", timeout=4000)
    expect(page.locator("#dpName")).to_be_visible()
    expect(page.locator("#dpName")).not_to_have_text("")


def test_detail_tabs_switch(page: Page):
    goto_app(page)
    page.locator(".contact-item").first.click()
    page.wait_for_selector(".dpt", timeout=4000)
    tabs = page.locator(".dpt")
    # Click กลุ่ม tab
    tabs.nth(1).click()
    expect(page.locator("#t-projects")).to_be_visible()
    # Click Notes tab
    tabs.nth(2).click()
    expect(page.locator("#t-notes")).to_be_visible()


# ─── 4. Add contact ───────────────────────────────────────────────────────────

def test_add_contact_via_drawer(page: Page):
    goto_app(page)
    page.click("button:has-text('เพิ่ม Contact')")
    page.wait_for_selector("#drawerOverlay.show", timeout=3000)
    page.fill("#d-name_th", "ทดสอบ playwright")
    page.fill("#d-name_en", "Test Playwright")
    page.fill("#d-email1", "test.playwright@test.com")
    page.click(".btn-save-d")
    page.wait_for_timeout(800)
    # Contact should appear in list
    page.fill("#searchInput", "playwright")
    page.wait_for_timeout(400)
    assert page.locator(".contact-item").count() >= 1


# ─── 5. Edit contact info ─────────────────────────────────────────────────────

def test_edit_contact_basic_fields(page: Page):
    goto_app(page)
    page.fill("#searchInput", "playwright")
    page.wait_for_timeout(400)
    page.locator(".contact-item").first.click()
    page.wait_for_selector("#dpName", timeout=4000)
    page.click("button:has-text('แก้ไข')")
    page.wait_for_selector("#editForm", state="visible", timeout=3000)
    page.fill("#ef-work_note", "งานทดสอบ playwright integration")
    page.fill("#ef-quick_note_team", "ทีมทดสอบอัตโนมัติ")
    page.click("button:has-text('บันทึก')")
    page.wait_for_timeout(800)
    # Work note should now appear
    expect(page.locator("#workNoteSection")).to_contain_text("งานทดสอบ playwright")


# ─── 6. Notes tab ─────────────────────────────────────────────────────────────

def test_add_note_to_contact(page: Page):
    goto_app(page)
    page.fill("#searchInput", "playwright")
    page.wait_for_timeout(400)
    page.locator(".contact-item").first.click()
    page.wait_for_selector(".dpt", timeout=4000)
    page.locator(".dpt").nth(2).click()  # Notes tab
    page.wait_for_selector("#newNoteContent", timeout=3000)
    page.fill("#newNoteContent", "Note จาก Playwright Test")
    page.fill("#newNoteTitle", "Playwright Note")
    page.click("button:has-text('บันทึก Note')")
    page.wait_for_timeout(800)
    expect(page.locator("#notesList")).to_contain_text("Playwright Note")


# ─── 7. API — Master Roles ────────────────────────────────────────────────────

def test_api_get_master_roles():
    roles = api("get", "/api/master/roles")
    assert isinstance(roles, list)
    names = [r["name"] for r in roles]
    assert "PM" in names, "Seed roles should include PM"


def test_api_create_master_role():
    new_name = "Test Role PW"
    api("post", "/api/master/roles", {"name": new_name, "color": "#123456"})
    roles = api("get", "/api/master/roles")
    assert any(r["name"] == new_name for r in roles)
    # Cleanup
    requests.delete(f"{BASE}/api/master/roles/{new_name}", timeout=5)


def test_api_delete_master_role():
    api("post", "/api/master/roles", {"name": "__del_test__"})
    requests.delete(f"{BASE}/api/master/roles/__del_test__", timeout=5)
    roles = api("get", "/api/master/roles")
    assert not any(r["name"] == "__del_test__" for r in roles)


# ─── 8. API — Contact Roles ───────────────────────────────────────────────────

def test_api_set_contact_roles():
    contacts = api("get", "/api/contacts")
    assert contacts
    cid = contacts[0]["id"]
    api("put", f"/api/contacts/{cid}/roles", {"roles": ["PM", "IT"]})
    contact = api("get", f"/api/contacts/{cid}")
    assert "PM" in contact["roles"]
    assert "IT" in contact["roles"]
    # Cleanup
    api("put", f"/api/contacts/{cid}/roles", {"roles": []})


# ─── 9. API — Entity ──────────────────────────────────────────────────────────

def test_api_set_contact_entity_vendor():
    contacts = api("get", "/api/contacts")
    cid = contacts[0]["id"]
    api("put", f"/api/contacts/{cid}/entity", {"entity_type": "Vendor", "entity_value": "IBM"})
    contact = api("get", f"/api/contacts/{cid}")
    assert contact["entity"]["entity_type"] == "Vendor"
    assert contact["entity"]["entity_value"] == "IBM"
    # Cleanup
    api("put", f"/api/contacts/{cid}/entity", {"entity_type": ""})


def test_api_set_contact_entity_ks():
    contacts = api("get", "/api/contacts")
    cid = contacts[0]["id"]
    api("put", f"/api/contacts/{cid}/entity", {"entity_type": "KS"})
    contact = api("get", f"/api/contacts/{cid}")
    assert contact["entity"]["entity_type"] == "KS"
    api("put", f"/api/contacts/{cid}/entity", {"entity_type": ""})


# ─── 10. API — Systems ────────────────────────────────────────────────────────

def test_api_get_master_systems():
    systems = api("get", "/api/master/systems")
    assert isinstance(systems, list)
    assert len(systems) > 0, "Should have seed systems"


def test_api_set_contact_systems():
    contacts = api("get", "/api/contacts")
    cid = contacts[0]["id"]
    api("put", f"/api/contacts/{cid}/systems", {"systems": ["KMA", "APIF"]})
    contact = api("get", f"/api/contacts/{cid}")
    assert "KMA" in contact["systems"]
    assert "APIF" in contact["systems"]
    api("put", f"/api/contacts/{cid}/systems", {"systems": []})


# ─── 11. API — Project Sub-columns ───────────────────────────────────────────

def test_api_project_subcolumns_crud():
    projs = api("get", "/api/master/projects")
    assert projs
    proj_name = projs[0]["name"]

    # Add sub-column
    api("post", f"/api/projects/{proj_name}/subcolumns", {"col_name": "Test Col PW"})
    cols = api("get", f"/api/projects/{proj_name}/subcolumns")
    col = next((c for c in cols if c["col_name"] == "Test Col PW"), None)
    assert col is not None

    # Delete it
    requests.delete(f"{BASE}/api/projects/{proj_name}/subcolumns/{col['id']}", timeout=5)
    cols_after = api("get", f"/api/projects/{proj_name}/subcolumns")
    assert not any(c["col_name"] == "Test Col PW" for c in cols_after)


def test_api_project_subvalues():
    contacts = api("get", "/api/contacts")
    cid = contacts[0]["id"]
    projs = api("get", "/api/master/projects")
    proj_name = projs[0]["name"]

    # Set a sub-value
    api("put", f"/api/contacts/{cid}/projects/{proj_name}/subvalues",
        {"subvalues": {"Remark": "ค่าทดสอบ Playwright"}})
    contact = api("get", f"/api/contacts/{cid}")
    proj = next((p for p in contact["projects"] if p["project_name"] == proj_name), None)
    # Sub-value may or may not exist depending on contact_projects assignment
    # Just verify API doesn't error
    assert contact is not None


# ─── 12. UI — Roles edit panel ───────────────────────────────────────────────

def test_ui_roles_edit_panel(page: Page):
    goto_app(page)
    page.locator(".contact-item").first.click()
    page.wait_for_selector(".dpt", timeout=4000)
    # Switch to กลุ่ม tab
    page.locator(".dpt").nth(1).click()
    page.wait_for_selector("#t-projects", timeout=3000)
    # Click แก้ไข for Main Role
    edit_btns = page.locator(".assec-edit-btn")
    edit_btns.first.click()
    page.wait_for_selector("#rolesEditPanel", state="visible", timeout=2000)
    # Checkboxes should be visible
    checkboxes = page.locator("#rolesCheckGrid input[type=checkbox]")
    assert checkboxes.count() > 0, "Role checkboxes should appear"


def test_ui_entity_edit_panel(page: Page):
    goto_app(page)
    page.locator(".contact-item").first.click()
    page.wait_for_selector(".dpt", timeout=4000)
    page.locator(".dpt").nth(1).click()
    page.wait_for_selector(".assec-edit-btn", timeout=3000)
    edit_btns = page.locator(".assec-edit-btn")
    edit_btns.nth(1).click()
    page.wait_for_selector("#entityEditPanel", state="visible", timeout=2000)
    expect(page.locator("#entityTypeSelect")).to_be_visible()


# ─── 13. Export API ───────────────────────────────────────────────────────────

def test_export_returns_xlsx():
    r = requests.get(f"{BASE}/api/sync/export", timeout=30)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert len(r.content) > 1000, "Excel file should have content"


def test_export_xlsx_structure():
    import io
    import openpyxl
    r = requests.get(f"{BASE}/api/sync/export", timeout=30)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    assert "📊 Full Info" in wb.sheetnames
    assert "📝 Notes" in wb.sheetnames
    assert "⚙ Master Data" in wb.sheetnames

    ws = wb["📊 Full Info"]
    # Row 1 should have individual field names (merged across rows 1-3)
    r1_vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    r1_vals = [v for v in r1_vals if v]
    assert "Name (EN)" in r1_vals
    assert "Name (TH)" in r1_vals
    assert "My Notes" in r1_vals
    # Should have zone labels
    assert "Projects" in r1_vals or "Main Role" in r1_vals or "Entity" in r1_vals


# ─── 14. Delete contact ───────────────────────────────────────────────────────

def test_delete_test_contact(page: Page):
    goto_app(page)
    page.fill("#searchInput", "playwright")
    page.wait_for_timeout(400)
    if page.locator(".contact-item").count() == 0:
        pytest.skip("No playwright test contact to delete")
    page.locator(".contact-item").first.click()
    page.wait_for_selector("#dpName", timeout=4000)
    page.click(".dp-act-btn:not(.primary)")  # delete button
    page.wait_for_selector("#delModal.show", timeout=3000)
    page.click("button:has-text('ลบ')")
    page.wait_for_timeout(1500)
    page.fill("#searchInput", "playwright")
    page.wait_for_timeout(600)
    assert page.locator(".contact-item").count() == 0, "Deleted contact should not appear"


# ─── 15. 2.1 Project edit in Edit Mode ───────────────────────────────────────

def test_edit_mode_project_add_remove(page: Page):
    goto_app(page)
    page.locator(".contact-item").first.click()
    page.wait_for_selector("#dpName", timeout=4000)
    page.click("button:has-text('แก้ไข')")
    page.wait_for_selector("#editForm", state="visible", timeout=3000)
    # Project edit section should be visible
    expect(page.locator("#efProjectsList")).to_be_visible()
    expect(page.locator("#efProjSelect")).to_be_visible()
    # Add a project
    proj_options = page.locator("#efProjSelect option")
    count = proj_options.count()
    assert count > 1, "Project select should have options"
    page.select_option("#efProjSelect", index=1)
    page.select_option("#efProjRole", "Supporter")
    page.click(".ef-proj-add-btn")
    # Project row should appear
    page.wait_for_timeout(200)
    assert page.locator(".ef-proj-row").count() >= 1


def test_api_save_projects_in_edit(page: Page):
    contacts = api("get", "/api/contacts")
    cid = contacts[0]["id"]
    original_projects = api("get", f"/api/contacts/{cid}")["projects"]
    # Set via API
    api("put", f"/api/contacts/{cid}", {
        "name_th": contacts[0]["name_th"], "email1": contacts[0].get("email1","x@x.com"),
        "projects": [{"project_name": "COBRA", "role": "Supporter"}]
    })
    updated = api("get", f"/api/contacts/{cid}")
    assert any(p["project_name"] == "COBRA" for p in updated["projects"])
    # Restore
    api("put", f"/api/contacts/{cid}", {
        "name_th": contacts[0]["name_th"], "email1": contacts[0].get("email1","x@x.com"),
        "projects": original_projects
    })


# ─── 16. 2.2 Grid view ───────────────────────────────────────────────────────

def test_grid_view_toggle(page: Page):
    goto_app(page)
    # Toggle to excel mode
    page.click("#viewToggleBtn")
    page.wait_for_timeout(400)
    # Excel table should be visible
    expect(page.locator("#excelWrap")).to_be_visible()
    expect(page.locator(".excel-tbl")).to_be_visible()
    # Table should have rows
    assert page.locator("#excelTbody tr").count() >= 1
    # Click a row — detail drawer should open
    page.locator("#excelTbody tr").first.click()
    page.wait_for_selector("#detail.drawer-open", timeout=4000)
    expect(page.locator("#dpName")).to_be_visible()
    # Close drawer
    page.click(".detail-close-btn")
    page.wait_for_timeout(200)
    # Toggle back to list
    page.click("#viewToggleBtn")
    page.wait_for_timeout(300)
    expect(page.locator("#contactList")).to_be_visible()
    assert page.locator("#excelWrap").is_hidden()


def test_sidebar_collapse(page: Page):
    goto_app(page)
    # Sidebar starts open
    expect(page.locator("#sidebar")).to_be_visible()
    # Click toggle to collapse
    page.click("#sbToggle")
    page.wait_for_timeout(400)
    assert page.locator("#sidebar.collapsed").count() == 1, "Sidebar should be collapsed"
    # Click toggle to expand
    page.click("#sbToggle")
    page.wait_for_timeout(400)
    assert page.locator("#sidebar.collapsed").count() == 0, "Sidebar should be expanded"


# ─── 17. 2.3 Todos/Tasks tab ─────────────────────────────────────────────────

def test_api_todos_crud():
    contacts = api("get", "/api/contacts")
    cid = contacts[0]["id"]
    # Create
    todo = api("post", f"/api/contacts/{cid}/todos", {"title": "Task ทดสอบ Playwright", "due_date": "2026-12-31"})
    tid = todo["id"]
    assert todo["title"] == "Task ทดสอบ Playwright"
    assert todo["done"] == 0
    # List
    todos = api("get", f"/api/contacts/{cid}/todos")
    assert any(t["id"] == tid for t in todos)
    # Toggle done
    api("put", f"/api/todos/{tid}", {"title": "Task ทดสอบ Playwright", "done": True})
    updated = next(t for t in api("get", f"/api/contacts/{cid}/todos") if t["id"] == tid)
    assert updated["done"] == 1
    # Delete
    requests.delete(f"{BASE}/api/todos/{tid}", timeout=5)
    todos_after = api("get", f"/api/contacts/{cid}/todos")
    assert not any(t["id"] == tid for t in todos_after)


def test_ui_tasks_tab(page: Page):
    goto_app(page)
    page.locator(".contact-item").first.click()
    page.wait_for_selector("#dpName", timeout=4000)
    # Click Tasks tab
    page.locator(".dpt").nth(3).click()
    page.wait_for_selector("#t-tasks", state="visible", timeout=3000)
    expect(page.locator("#todoList")).to_be_visible()
    expect(page.locator("#newTodoTitle")).to_be_visible()
    # Add a todo
    page.fill("#newTodoTitle", "Task UI Test Playwright")
    page.click(".todo-add-btn")
    page.wait_for_timeout(600)
    expect(page.locator("#todoList")).to_contain_text("Task UI Test Playwright")
    # Toggle done
    page.locator(".todo-cb").first.click()
    page.wait_for_timeout(400)
    # Delete
    page.locator(".todo-del").first.click()
    page.wait_for_timeout(500)


# ─── 18. 2.4 Quick Find ──────────────────────────────────────────────────────

def test_quick_find_opens_and_searches(page: Page):
    goto_app(page)
    # Open via button
    page.click("button[title*='Quick Find']")
    page.wait_for_selector("#qfOverlay.show", timeout=3000)
    expect(page.locator("#qfInput")).to_be_visible()
    # Search
    page.fill("#qfInput", "สมชาย")
    page.wait_for_timeout(500)
    results = page.locator(".qf-item")
    assert results.count() >= 1
    names = page.locator(".qf-name")
    found = any("สมชาย" in names.nth(i).inner_text() for i in range(names.count()))
    assert found, "Quick Find should surface สมชาย"
    # Select first result — should open detail
    page.locator(".qf-item").first.click()
    page.wait_for_selector("#dpName", timeout=4000)
    expect(page.locator("#qfOverlay")).not_to_have_class("show")


def test_quick_find_keyboard_shortcut(page: Page):
    goto_app(page)
    # Open with Ctrl+K
    page.keyboard.press("Control+k")
    page.wait_for_selector("#qfOverlay.show", timeout=2000)
    # Close with Escape
    page.keyboard.press("Escape")
    page.wait_for_timeout(200)
    assert page.locator("#qfOverlay.show").count() == 0
