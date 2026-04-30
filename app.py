# MyContacts v3.4 — Contact photo upload: POST/DELETE /api/contacts/<id>/photo; auto-resize 400×400 via Pillow
# MyContacts v3.3 — Master data reorder endpoints; sort_order in projects/tags APIs
# MyContacts v3.1 — B+C Import: full-zone Excel parse, auto-create master data, new-master banner
# v3.0 — Sub-col values in contact edit; Import preview/execute/clear; sync disabled
# v2.3: Master Data management modal (Projects, Roles, Systems, Tags CRUD)
# v2.2: Added /api/sync/export_view endpoint for custom-column Excel export from grid view
# v2.1: Grid view, Todos/Tasks tab, Quick Find (Ctrl+K), Project edit in Edit Mode
# v2.0: New schema — work_note, quick_note_team, master_roles/contact_roles, contact_entity,
#        master_systems/contact_systems; project_subcolumns + contact_project_subvalues
# v1.x: CRUD, Excel sync, Full Info sheet, master_apps, Single Team
#
# Run: pip install flask pandas openpyxl
#      python app.py  → open http://localhost:5000

from flask import Flask, jsonify, request, render_template, send_file
import sqlite3, os, json, threading, time
from datetime import datetime
import io

try:
    from PIL import Image as _PILImage
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

app = Flask(__name__)
DB          = os.path.join(os.path.dirname(__file__), 'contacts.db')
EXCEL       = os.path.join(os.path.dirname(__file__), 'mycontacts.xlsx')
AVATARS_DIR = os.path.join(os.path.dirname(__file__), 'static', 'avatars')
os.makedirs(AVATARS_DIR, exist_ok=True)

# ─── DB helpers ───────────────────────────────────────────────────────────────

def db_conn():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = db_conn()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS contacts (
            id               TEXT PRIMARY KEY,
            name_th          TEXT NOT NULL,
            name_en          TEXT,
            nickname         TEXT,
            team             TEXT,
            sub_team         TEXT,
            org_role         TEXT,
            direct_report    TEXT,
            email1           TEXT,
            email2           TEXT,
            phone            TEXT,
            line_id          TEXT,
            note_short       TEXT,
            general_note     TEXT,
            work_note        TEXT,
            quick_note_team  TEXT,
            associated_to_meeting TEXT,
            created_at       TEXT DEFAULT (datetime('now','localtime')),
            updated_at       TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS contact_projects (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id   TEXT NOT NULL,
            project_name TEXT NOT NULL,
            role         TEXT DEFAULT 'Supporter',
            note         TEXT,
            sort_order   INTEGER DEFAULT 0,
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS notes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id TEXT NOT NULL,
            note_date  TEXT,
            title      TEXT,
            content    TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS tags (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT UNIQUE NOT NULL,
            color      TEXT DEFAULT '#64748B',
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS contact_tags (
            contact_id TEXT NOT NULL,
            tag_id     INTEGER NOT NULL,
            PRIMARY KEY (contact_id, tag_id),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE,
            FOREIGN KEY (tag_id)     REFERENCES tags(id)     ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS master_projects (
            name       TEXT PRIMARY KEY,
            color      TEXT DEFAULT '#64748B',
            short_name TEXT,
            sort_order INTEGER DEFAULT 0,
            visible    INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS master_teams (
            name  TEXT PRIMARY KEY,
            color TEXT DEFAULT '#64748B'
        );
        -- v2.0: Main Role (multi-select, user-managed)
        CREATE TABLE IF NOT EXISTS master_roles (
            name       TEXT PRIMARY KEY,
            color      TEXT DEFAULT '#64748B',
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS contact_roles (
            contact_id TEXT NOT NULL,
            role_name  TEXT NOT NULL,
            PRIMARY KEY (contact_id, role_name),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        -- v2.0: Entity (1 per contact: KS / Vendor / Subsidiary / 3rd Party / Other Bank)
        CREATE TABLE IF NOT EXISTS contact_entity (
            contact_id   TEXT PRIMARY KEY,
            entity_type  TEXT NOT NULL,
            entity_value TEXT,
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        -- v2.0: Related Systems/Areas (Y/N per system, replaces contact_apps+role)
        CREATE TABLE IF NOT EXISTS master_systems (
            name       TEXT PRIMARY KEY,
            sort_order INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS contact_systems (
            contact_id  TEXT NOT NULL,
            system_name TEXT NOT NULL,
            PRIMARY KEY (contact_id, system_name),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        -- v2.0: Project sub-columns (per-project config)
        CREATE TABLE IF NOT EXISTS project_subcolumns (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT NOT NULL,
            col_name     TEXT NOT NULL,
            sort_order   INTEGER DEFAULT 0,
            visible      INTEGER DEFAULT 1,
            UNIQUE (project_name, col_name)
        );
        CREATE TABLE IF NOT EXISTS contact_project_subvalues (
            contact_id   TEXT NOT NULL,
            project_name TEXT NOT NULL,
            col_name     TEXT NOT NULL,
            value        TEXT,
            PRIMARY KEY (contact_id, project_name, col_name),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        -- v2.1: Todos per contact
        CREATE TABLE IF NOT EXISTS todos (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id TEXT NOT NULL,
            title      TEXT NOT NULL,
            done       INTEGER DEFAULT 0,
            due_date   TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        -- Legacy tables kept for compatibility
        CREATE TABLE IF NOT EXISTS master_apps (
            name  TEXT PRIMARY KEY,
            color TEXT DEFAULT '#6366F1'
        );
        CREATE TABLE IF NOT EXISTS contact_apps (
            contact_id TEXT NOT NULL,
            app_name   TEXT NOT NULL,
            role       TEXT DEFAULT 'Y',
            PRIMARY KEY (contact_id, app_name),
            FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        );
        CREATE TABLE IF NOT EXISTS sync_log (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            synced_at       TEXT,
            excel_path      TEXT,
            from_excel      INTEGER DEFAULT 0,
            to_excel        INTEGER DEFAULT 0,
            new_in_excel    INTEGER DEFAULT 0,
            new_in_db       INTEGER DEFAULT 0,
            notes_appended  INTEGER DEFAULT 0,
            projects_synced INTEGER DEFAULT 0,
            conflicts       TEXT DEFAULT '[]',
            errors          TEXT DEFAULT '[]'
        );
    """)
    conn.commit()

    # ── Migrate existing DBs ──────────────────────────────────────────────────
    cols = [r[1] for r in conn.execute("PRAGMA table_info(contacts)").fetchall()]
    for col, typedef in [('sub_team','TEXT'), ('work_note','TEXT'), ('quick_note_team','TEXT'),
                         ('associated_to_meeting','TEXT'), ('photo_path','TEXT')]:
        if col not in cols:
            conn.execute(f"ALTER TABLE contacts ADD COLUMN {col} {typedef}")
    proj_cols = [r[1] for r in conn.execute("PRAGMA table_info(contact_projects)").fetchall()]
    if 'sort_order' not in proj_cols:
        conn.execute("ALTER TABLE contact_projects ADD COLUMN sort_order INTEGER DEFAULT 0")
    mp_cols = [r[1] for r in conn.execute("PRAGMA table_info(master_projects)").fetchall()]
    if 'sort_order' not in mp_cols:
        conn.execute("ALTER TABLE master_projects ADD COLUMN sort_order INTEGER DEFAULT 0")
    if 'visible' not in mp_cols:
        conn.execute("ALTER TABLE master_projects ADD COLUMN visible INTEGER DEFAULT 1")
    tag_cols = [r[1] for r in conn.execute("PRAGMA table_info(tags)").fetchall()]
    if 'sort_order' not in tag_cols:
        conn.execute("ALTER TABLE tags ADD COLUMN sort_order INTEGER DEFAULT 0")
    conn.commit()

    # Migrate master_apps → master_systems (one-time)
    if conn.execute("SELECT COUNT(*) FROM master_systems").fetchone()[0] == 0:
        rows = conn.execute("SELECT name FROM master_apps").fetchall()
        if rows:
            conn.executemany("INSERT OR IGNORE INTO master_systems (name) VALUES (?)",
                             [(r['name'],) for r in rows])
    # Migrate contact_apps → contact_systems (one-time)
    if conn.execute("SELECT COUNT(*) FROM contact_systems").fetchone()[0] == 0:
        rows = conn.execute("SELECT contact_id, app_name FROM contact_apps").fetchall()
        if rows:
            conn.executemany("INSERT OR IGNORE INTO contact_systems (contact_id, system_name) VALUES (?,?)",
                             [(r['contact_id'], r['app_name']) for r in rows])
    conn.commit()

    c = conn.cursor()
    if c.execute("SELECT COUNT(*) FROM master_projects").fetchone()[0] == 0:
        _seed_master(conn)
    if c.execute("SELECT COUNT(*) FROM master_systems").fetchone()[0] == 0:
        _seed_systems(conn)
    if c.execute("SELECT COUNT(*) FROM master_roles").fetchone()[0] == 0:
        _seed_roles(conn)
    if c.execute("SELECT COUNT(*) FROM contacts").fetchone()[0] == 0:
        _seed_contacts(conn)
    conn.close()

def _seed_master(conn):
    conn.executemany("INSERT OR IGNORE INTO master_projects (name,color,short_name) VALUES (?,?,?)", [
        ('COBRA',   '#3B82F6', 'Core Banking'),
        ('ATLAS',   '#10B981', 'Settlement'),
        ('NEXUS',   '#F59E0B', 'Nexus'),
        ('PHOENIX', '#EF4444', 'Phoenix'),
    ])
    conn.executemany("INSERT OR IGNORE INTO master_teams VALUES (?,?)", [
        ('Core Banking IT',    '#6366F1'),
        ('Digital Banking IT', '#0EA5E9'),
        ('Settlement IT',      '#14B8A6'),
        ('BizOps',             '#8B5CF6'),
        ('PMO',                '#F97316'),
        ('Senior Management',  '#BE185D'),
        ('Vendor',             '#64748B'),
    ])
    conn.executemany("INSERT OR IGNORE INTO tags (name, color) VALUES (?,?)", [
        ('Management',    '#BE185D'),
        ('Vendor',        '#64748B'),
        ('Agile Champion','#10B981'),
        ('QA Expert',     '#3B82F6'),
    ])
    conn.commit()

def _seed_systems(conn):
    systems = ['KMA','ITMX','O9','APIF','RnT','KDC','GL','IP Contra','TLM','OFSAA',
               'BOTDMS','Actimize','KYC','AMLO','EDM','PMH','KSA','OneApp','APIM','DAP','AF1','KBOL Web','KBOL App']
    conn.executemany("INSERT OR IGNORE INTO master_systems (name) VALUES (?)", [(s,) for s in systems])
    conn.commit()

def _seed_roles(conn):
    roles = [('PM','#3B82F6',0),('Biz, PO','#10B981',1),('IT','#6366F1',2),
             ('User','#F59E0B',3),('QA','#EF4444',4)]
    conn.executemany("INSERT OR IGNORE INTO master_roles (name,color,sort_order) VALUES (?,?,?)", roles)
    conn.commit()

def _seed_apps(conn):
    apps = ['PMH','KSA','OneApp','APIF','APIM','DAP','OFSAA','GL','AF1','KBOL Web','KBOL App']
    conn.executemany("INSERT OR IGNORE INTO master_apps (name) VALUES (?)", [(a,) for a in apps])
    conn.commit()

def _seed_contacts(conn):
    contacts = [
        ('C001','สมชาย มีสุข','Somchai Meesuk','ชาย','Core Banking IT','QA Lead','นายสมบัติ วงษ์วิทยา','somchai.m@company.co.th',None,'081-234-5678','@cha_somchai','เชี่ยวชาญ UAT Banking','ทำงานละเอียด ตรงต่อเวลา เหมาะกับงาน QA'),
        ('C002','วรรณา ดีใจ','Wanna Deejai','แนน','Digital Banking IT','Business Analyst','นางสาว อรุณี บุญมา','wanna.d@company.co.th',None,'082-345-6789','@nan_wanna','BA ผู้เชี่ยวชาญ Mobile','ประสบการณ์ Mobile Banking มากกว่า 5 ปี'),
        ('C003','ประสิทธิ์ มั่นคง','Prasit Mankong','ต้น','PMO','Project Manager','นาย วิชัย ดีมา','prasit.m@company.co.th',None,'083-456-7890','@ton_prasit','PM ผู้เชี่ยวชาญ Agile',None),
        ('C004','นิตยา สุขสม','Nittaya Suksom','ตา','Settlement IT','System Analyst','นาย กิตติ พิมล','nittaya.s@company.co.th',None,'084-567-8901','@ta_nittaya','SA ด้าน Settlement',None),
        ('C005','อรุณี บุญมา','Arunee Boonma','อ้อ','Senior Management','VP Technology',None,'arunee.b@company.co.th',None,'085-678-9012',None,'VP สายงานเทคโนโลยี',None),
    ]
    conn.executemany("""INSERT INTO contacts
        (id,name_th,name_en,nickname,team,org_role,direct_report,email1,email2,phone,line_id,note_short,general_note)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""", contacts)

    conn.executemany("INSERT INTO contact_projects (contact_id,project_name,role,note) VALUES (?,?,?,?)", [
        ('C001','COBRA','Main Impact',None), ('C001','NEXUS','Supporter',None),
        ('C002','COBRA','Supporter',None),   ('C002','ATLAS','Main Impact',None),
        ('C003','COBRA','Project Manager',None), ('C003','PHOENIX','Project Manager',None),
        ('C004','ATLAS','Main Impact',None), ('C004','NEXUS','Main Impact',None),
        ('C005','COBRA','Supporter','Executive Sponsor'),
        ('C005','ATLAS','Supporter','Executive Sponsor'),
    ])

    conn.executemany("INSERT INTO notes (contact_id,note_date,title,content) VALUES (?,?,?,?)", [
        ('C001','2026-04-20','ประชุม UAT Planning','หารือแผน UAT สำหรับ COBRA Phase 2 กำหนด test case 280 รายการ'),
        ('C001','2026-04-15','Follow-up Bug Report','Bug #2341 แก้ไขแล้ว รอ Retest ในวันศุกร์'),
        ('C002','2026-04-22','BA Workshop','นำเสนอ requirements ใหม่สำหรับ Mobile App v3.0'),
        ('C003','2026-04-18','Status Report','COBRA Phase 2 progress 65% on track สำหรับ deadline มิถุนายน'),
    ])

    conn.executemany("INSERT OR IGNORE INTO contact_tags VALUES (?,?)", [
        ('C001',4), ('C002',3), ('C003',1), ('C005',1),
    ])
    conn.commit()

def _next_id(conn):
    row = conn.execute("SELECT id FROM contacts ORDER BY id DESC LIMIT 1").fetchone()
    if not row:
        return 'C001'
    return f"C{int(row['id'][1:])+1:03d}"

def _contact_dict(row, conn, full=False):
    d = dict(row)
    cid = d['id']
    # Projects with sub-values
    proj_rows = conn.execute(
        "SELECT * FROM contact_projects WHERE contact_id=? ORDER BY sort_order, id", (cid,)).fetchall()
    projects = []
    for p in proj_rows:
        pd_ = dict(p)
        subvals = conn.execute(
            "SELECT col_name, value FROM contact_project_subvalues WHERE contact_id=? AND project_name=?",
            (cid, p['project_name'])).fetchall()
        pd_['subvalues'] = {r['col_name']: r['value'] for r in subvals}
        projects.append(pd_)
    d['projects'] = projects
    d['tags'] = [dict(r) for r in conn.execute(
        "SELECT t.id,t.name,t.color FROM tags t JOIN contact_tags ct ON ct.tag_id=t.id WHERE ct.contact_id=?", (cid,))]
    # v2.0 fields
    d['roles'] = [r['role_name'] for r in conn.execute(
        "SELECT role_name FROM contact_roles WHERE contact_id=? ORDER BY role_name", (cid,)).fetchall()]
    entity = conn.execute("SELECT entity_type, entity_value FROM contact_entity WHERE contact_id=?", (cid,)).fetchone()
    d['entity'] = dict(entity) if entity else None
    d['systems'] = [r['system_name'] for r in conn.execute(
        "SELECT system_name FROM contact_systems WHERE contact_id=? ORDER BY system_name", (cid,)).fetchall()]
    # Legacy apps kept for backward compat
    d['apps'] = [dict(r) for r in conn.execute(
        "SELECT app_name, role FROM contact_apps WHERE contact_id=? ORDER BY app_name", (cid,))]
    d['note_count'] = conn.execute("SELECT COUNT(*) FROM notes WHERE contact_id=?", (cid,)).fetchone()[0]
    d['todo_count'] = conn.execute("SELECT COUNT(*) FROM todos WHERE contact_id=? AND done=0", (cid,)).fetchone()[0]
    # Photo URL with cache-bust timestamp
    _fp = d.get('photo_path')
    if _fp:
        _abs = os.path.join(AVATARS_DIR, _fp)
        _ts  = int(os.path.getmtime(_abs)) if os.path.exists(_abs) else 0
        d['photo_url'] = f'/static/avatars/{_fp}?v={_ts}'
    else:
        d['photo_url'] = None
    if full:
        d['notes'] = [dict(r) for r in conn.execute(
            "SELECT * FROM notes WHERE contact_id=? ORDER BY note_date DESC, id DESC", (cid,))]
        d['todos'] = [dict(r) for r in conn.execute(
            "SELECT * FROM todos WHERE contact_id=? ORDER BY done, due_date, id", (cid,))]
    return d

# ─── Excel helpers ─────────────────────────────────────────────────────────────

def _str(row, key, default=''):
    """Safe string extraction from pandas row — strips nan/None."""
    v = row.get(key, default)
    s = str(v).strip()
    return '' if s in ('nan', 'None', 'NaT', 'NaN') else s

def _parse_dt(s):
    """Parse datetime string → datetime object, return None if unparseable."""
    if not s or str(s).strip() in ('nan', 'None', 'NaT', 'NaN', ''):
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(s)[:19], fmt)
        except ValueError:
            pass
    return None

CONTACT_FIELDS = [
    'name_th','name_en','nickname','team','sub_team','org_role','direct_report',
    'email1','email2','phone','line_id','note_short','general_note',
    'associated_to_meeting',
]

CONTACT_COL_MAP = {
    'ContactID':'id', 'ชื่อ (ไทย)':'name_th', 'Name (Eng)':'name_en',
    'ชื่อเล่น':'nickname', 'Team':'team', 'Sub-Team':'sub_team',
    'ตำแหน่ง':'org_role', 'Direct Report':'direct_report',
    'Email (หลัก)':'email1', 'Email (2)':'email2',
    'โทรศัพท์':'phone', 'Line ID':'line_id',
    'หมายเหตุสั้น':'note_short', 'General Note':'general_note',
    'Updated':'excel_updated',
}

def _build_fullinfo_sheet(ws, conn):
    """
    Export matching template structure (3-row header):
      Row 1: Individual field names (merged R1:R3) | zone labels (Projects/Roles/Entity/Systems/Tags) | Note1-10 (merged R1:R3)
      Row 2: blank for indiv | project names (merged across sub-cols) | role names | entity types | systems | tags
      Row 3: blank for indiv | sub-column names per project | blank for others
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Master data ──────────────────────────────────────────────────────────
    projs     = conn.execute("SELECT name FROM master_projects WHERE visible=1 ORDER BY sort_order,name").fetchall()
    roles_m   = conn.execute("SELECT name,color FROM master_roles ORDER BY sort_order,name").fetchall()
    ent_types = ['KS','Vendor','Subsidiary','3rd Party','Other Bank']
    systems   = conn.execute("SELECT name FROM master_systems ORDER BY sort_order,name").fetchall()
    tags      = conn.execute("SELECT name FROM tags ORDER BY sort_order,name").fetchall()
    contacts  = conn.execute("SELECT * FROM contacts ORDER BY id").fetchall()

    proj_list   = [r['name'] for r in projs]
    role_list   = [r['name'] for r in roles_m]
    role_colors = {r['name']: r['color'] for r in roles_m}
    sys_list    = [r['name'] for r in systems]
    tag_list    = [r['name'] for r in tags]

    # Sub-columns per project: {proj_name: [col_name, ...]}
    proj_subcols = {}
    for pname in proj_list:
        cols = conn.execute(
            "SELECT col_name FROM project_subcolumns WHERE project_name=? AND visible=1 ORDER BY sort_order,id",
            (pname,)).fetchall()
        if cols:
            proj_subcols[pname] = [r['col_name'] for r in cols]
        else:
            proj_subcols[pname] = ['Remark']  # default sub-column

    NOTE_COUNT = 10

    # ── Individual fields (merged rows 1-3) ───────────────────────────────────
    INDIV = [
        ('ContactID',               'id',               12),
        ('Name (EN)',               'name_en',          28),
        ('Name (TH)',               'name_th',          28),
        ('Email 1',                 'email1',            26),
        ('Email 2',                 'email2',            20),
        ('Direct Report to',        'direct_report',     20),
        ('Team',                    'team',              16),
        ('Sub-Team',                'sub_team',          14),
        ('Mobile',                  'phone',             14),
        ('Line Name',               'line_id',           14),
        ('My Notes',                'general_note',      30),
        ('Nickname',                'nickname',          12),
        ('Notes',                   'work_note',         30),
        ('Quick Notes about Team',  'quick_note_team',   24),
    ]
    NI = len(INDIV)   # number of individual columns

    # ── Zone definitions ─────────────────────────────────────────────────────
    # Each project occupies: 1 (flag) + len(subcols) columns
    def proj_width(pname): return 1 + len(proj_subcols.get(pname, ['Remark']))
    total_proj_cols = sum(proj_width(p) for p in proj_list)
    nr_roles   = len(role_list)
    nr_entity  = len(ent_types)
    nr_systems = len(sys_list)
    nr_tags    = len(tag_list)

    # Starting column (1-indexed) for each zone
    z_proj   = NI + 1
    z_roles  = z_proj  + total_proj_cols
    z_entity = z_roles + nr_roles
    z_sys    = z_entity + nr_entity
    z_tags   = z_sys   + nr_systems
    z_notes  = z_tags  + nr_tags
    total_cols = z_notes + NOTE_COUNT - 1

    # ── Style helpers ─────────────────────────────────────────────────────────
    def hdr_cell(row, col, value, bg, fg='1E293B', bold=True, size=9, wrap=False, h_align='center'):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.font = Font(bold=bold, size=size, color=fg, name='Calibri')
        cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=wrap)
        return cell

    def merge_if(r1, c1, r2, c2):
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    # Zone colors: (zone_fill_R1, col_fill_R2R3)
    Z_PROJ   = ('FEF3C7', 'FDE68A')
    Z_ROLE   = ('DCFCE7', '86EFAC')
    Z_ENT    = ('FFF7ED', 'FED7AA')
    Z_SYS    = ('EDE9FE', 'C4B5FD')
    Z_TAG    = ('FCE7F3', 'F9A8D4')
    Z_NOTE   = ('F1F5F9', 'CBD5E1')
    Z_INDIV  = ('DBEAFE', '93C5FD')

    # ── ROW 1: Zone labels + Individual field names ────────────────────────────
    # Individual fields: merge R1:R3 each
    for ci, (hdr, _, _) in enumerate(INDIV, 1):
        hdr_cell(1, ci, hdr, Z_INDIV[0], bold=True, size=9, wrap=True)
        merge_if(1, ci, 3, ci)

    # Projects zone label
    if total_proj_cols > 0:
        hdr_cell(1, z_proj, 'Projects', Z_PROJ[0], bold=True, size=10)
        merge_if(1, z_proj, 1, z_proj + total_proj_cols - 1)

    # Main Role zone label
    if nr_roles > 0:
        hdr_cell(1, z_roles, 'Main Role', Z_ROLE[0], bold=True, size=10)
        merge_if(1, z_roles, 1, z_roles + nr_roles - 1)

    # Entity zone label
    if nr_entity > 0:
        hdr_cell(1, z_entity, 'Entity', Z_ENT[0], bold=True, size=10)
        merge_if(1, z_entity, 1, z_entity + nr_entity - 1)

    # Related Systems zone label
    if nr_systems > 0:
        hdr_cell(1, z_sys, 'Related Systems/Areas', Z_SYS[0], bold=True, size=10)
        merge_if(1, z_sys, 1, z_sys + nr_systems - 1)

    # Custom Tags zone label
    if nr_tags > 0:
        hdr_cell(1, z_tags, 'Custom Tags', Z_TAG[0], bold=True, size=10)
        merge_if(1, z_tags, 1, z_tags + nr_tags - 1)

    # Note1-Note10 (merged R1:R3 each)
    for i in range(NOTE_COUNT):
        col = z_notes + i
        hdr_cell(1, col, f'Note{i+1}', Z_NOTE[0], bold=True, size=9)
        merge_if(1, col, 3, col)

    # ── ROW 2: Project names (merged across sub-cols), Roles, Entity types, Systems, Tags ──
    col = z_proj
    for pname in proj_list:
        w = proj_width(pname)
        hdr_cell(2, col, pname, Z_PROJ[1], bold=True, size=9, wrap=True)
        merge_if(2, col, 2, col + w - 1)
        col += w

    for rname in role_list:
        c = role_colors.get(rname, '64748B').lstrip('#')
        hdr_cell(2, z_roles + role_list.index(rname), rname, Z_ROLE[1], bold=True, size=9, wrap=True)
        merge_if(2, z_roles + role_list.index(rname), 3, z_roles + role_list.index(rname))

    for ei, etype in enumerate(ent_types):
        hdr_cell(2, z_entity + ei, etype, Z_ENT[1], bold=True, size=9, wrap=True)
        merge_if(2, z_entity + ei, 3, z_entity + ei)

    for si, sname in enumerate(sys_list):
        hdr_cell(2, z_sys + si, sname, Z_SYS[1], bold=True, size=9, wrap=True)
        merge_if(2, z_sys + si, 3, z_sys + si)

    for ti, tname in enumerate(tag_list):
        hdr_cell(2, z_tags + ti, tname, Z_TAG[1], bold=True, size=9, wrap=True)
        merge_if(2, z_tags + ti, 3, z_tags + ti)

    # ── ROW 3: Project sub-column names ──────────────────────────────────────
    col = z_proj
    for pname in proj_list:
        # flag col = project name again
        hdr_cell(3, col, pname, Z_PROJ[1], bold=False, size=8, wrap=True)
        col += 1
        for scname in proj_subcols.get(pname, ['Remark']):
            hdr_cell(3, col, scname, Z_PROJ[1], bold=False, size=8, wrap=True)
            col += 1

    # ── DATA ROWS ─────────────────────────────────────────────────────────────
    STRIPE = PatternFill('solid', fgColor='F8FAFC')
    for ri, c in enumerate(contacts, 4):
        cid = c['id']

        # Individual fields
        for ci, (_, field, _) in enumerate(INDIV, 1):
            ws.cell(row=ri, column=ci, value=c[field] or '')

        # ── Projects ──────────────────────────────────────────────────────────
        cp_flags = {r['project_name'] for r in conn.execute(
            "SELECT project_name FROM contact_projects WHERE contact_id=?", (cid,)).fetchall()}
        cp_notes = {r['project_name']: r['note'] for r in conn.execute(
            "SELECT project_name, note FROM contact_projects WHERE contact_id=?", (cid,)).fetchall()}
        subvals  = {}
        for r in conn.execute(
            "SELECT project_name, col_name, value FROM contact_project_subvalues WHERE contact_id=?",
            (cid,)).fetchall():
            subvals.setdefault(r['project_name'], {})[r['col_name']] = r['value']

        col = z_proj
        for pname in proj_list:
            in_proj = pname in cp_flags
            ws.cell(row=ri, column=col, value='Y' if in_proj else '')
            col += 1
            for scname in proj_subcols.get(pname, ['Remark']):
                if scname == 'Remark':
                    v = cp_notes.get(pname, '') if in_proj else ''
                else:
                    v = subvals.get(pname, {}).get(scname, '') if in_proj else ''
                ws.cell(row=ri, column=col, value=v or '')
                col += 1

        # ── Main Role (Y/N) ──────────────────────────────────────────────────
        c_roles = {r['role_name'] for r in conn.execute(
            "SELECT role_name FROM contact_roles WHERE contact_id=?", (cid,)).fetchall()}
        for ri2, rname in enumerate(role_list):
            ws.cell(row=ri, column=z_roles + ri2, value='Y' if rname in c_roles else '')

        # ── Entity ───────────────────────────────────────────────────────────
        ent_row = conn.execute(
            "SELECT entity_type, entity_value FROM contact_entity WHERE contact_id=?",
            (cid,)).fetchone()
        for ei, etype in enumerate(ent_types):
            if ent_row and ent_row['entity_type'] == etype:
                # For KS/Subsidiary/3rd Party: 'Y'; for Vendor/Other Bank: company name
                if etype in ('Vendor', 'Other Bank'):
                    v = ent_row['entity_value'] or 'Y'
                else:
                    v = 'Y'
            else:
                v = ''
            ws.cell(row=ri, column=z_entity + ei, value=v)

        # ── Systems (Y/N) ────────────────────────────────────────────────────
        c_sys = {r['system_name'] for r in conn.execute(
            "SELECT system_name FROM contact_systems WHERE contact_id=?", (cid,)).fetchall()}
        for si, sname in enumerate(sys_list):
            ws.cell(row=ri, column=z_sys + si, value='Y' if sname in c_sys else '')

        # ── Tags (Y/N) ───────────────────────────────────────────────────────
        c_tags = {r['name'] for r in conn.execute(
            "SELECT t.name FROM tags t JOIN contact_tags ct ON ct.tag_id=t.id WHERE ct.contact_id=?",
            (cid,)).fetchall()}
        for ti, tname in enumerate(tag_list):
            ws.cell(row=ri, column=z_tags + ti, value='Y' if tname in c_tags else '')

        # ── Notes (latest 10, formatted YYYY-MM-DD <Title>\n<Content>) ───────
        note_rows = conn.execute(
            "SELECT note_date, title, content FROM notes WHERE contact_id=? ORDER BY note_date DESC, id DESC LIMIT ?",
            (cid, NOTE_COUNT)).fetchall()
        for ni, nr in enumerate(note_rows):
            date_part = str(nr['note_date'] or '')[:10]
            title_part = f" {nr['title']}" if nr['title'] else ''
            txt = f"{date_part}{title_part}\n{nr['content']}"
            cell = ws.cell(row=ri, column=z_notes + ni, value=txt)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        if ri % 2 == 0:
            for ci in range(1, total_cols + 1):
                ws.cell(row=ri, column=ci).fill = STRIPE

    # ── Column widths ────────────────────────────────────────────────────────
    for ci, (_, _, w) in enumerate(INDIV, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    col = z_proj
    for pname in proj_list:
        ws.column_dimensions[get_column_letter(col)].width = 6   # flag
        col += 1
        for _ in proj_subcols.get(pname, ['Remark']):
            ws.column_dimensions[get_column_letter(col)].width = 18
            col += 1

    for ci in range(z_roles, z_notes):
        ws.column_dimensions[get_column_letter(ci)].width = 10

    for i in range(NOTE_COUNT):
        ws.column_dimensions[get_column_letter(z_notes + i)].width = 30

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = ws.cell(row=4, column=NI + 1)

def _build_notes_sheet(ws, conn):
    """Write Notes sheet (read-only export)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    hdrs = ['ContactID','ชื่อ','วันที่','หัวข้อ','เนื้อหา Note','Created']
    rows = conn.execute("""
        SELECT n.contact_id, c.name_th, n.note_date, n.title, n.content, n.created_at
        FROM notes n JOIN contacts c ON c.id=n.contact_id
        ORDER BY n.contact_id, n.note_date DESC""").fetchall()
    GREEN = PatternFill('solid', fgColor='1B5E20')
    HFONT = Font(bold=True, color='FFFFFF', name='Calibri')
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = GREEN; cell.font = HFONT
        cell.alignment = Alignment(horizontal='center')
    for ri, r in enumerate(rows, 2):
        for ci, v in enumerate(r, 1):
            ws.cell(row=ri, column=ci, value=v)
    widths = [10, 20, 12, 20, 50, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

def _build_master_sheet(ws, conn):
    """Write Master Data sheet: Projects, Teams, Roles, Systems, Tags in sections."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    BLUE  = PatternFill('solid', fgColor='1E3A5F')
    HFONT = Font(bold=True, color='FFFFFF', name='Calibri')
    SFONT = Font(bold=True, size=11, name='Calibri')

    def write_section(start_row, title, hdrs, rows):
        ws.cell(row=start_row, column=1, value=title).font = SFONT
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(row=start_row+1, column=ci, value=h)
            cell.fill = BLUE; cell.font = HFONT; cell.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(rows, start_row+2):
            for ci, v in enumerate(r, 1):
                ws.cell(row=ri, column=ci, value=v)
        return start_row + 2 + len(rows) + 1

    projs   = conn.execute("SELECT name, color, short_name FROM master_projects ORDER BY sort_order,name").fetchall()
    teams   = conn.execute("SELECT name, color FROM master_teams ORDER BY name").fetchall()
    roles   = conn.execute("SELECT name, color FROM master_roles ORDER BY sort_order,name").fetchall()
    systems = conn.execute("SELECT name FROM master_systems ORDER BY sort_order,name").fetchall()
    tags    = conn.execute("SELECT name, color FROM tags ORDER BY sort_order,name").fetchall()

    r = write_section(1,  'Projects',              ['Project Name','Color','Short Name'], [tuple(x) for x in projs])
    r = write_section(r,  'Teams',                 ['Team Name','Color'], [tuple(x) for x in teams])
    r = write_section(r,  'Main Roles',            ['Role Name','Color'], [tuple(x) for x in roles])
    r = write_section(r,  'Related Systems/Areas', ['System/App Name'], [(x['name'],) for x in systems])
    write_section(r, 'Custom Tags', ['Tag Name','Color'], [tuple(x) for x in tags])

    for ci, w in enumerate([28, 12, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def _parse_fullinfo_sheet(xl, conn, result, now_str, bias_s):
    """
    Parse Full Info sheet → update contacts, teams, apps, projects, tags in DB.
    Zone detection: forward-fill Row 0 labels (Contact Fields / Team / App / Projects / Tags).
    New column headers → auto-create master entities.
    Empty ContactID → generate new ID + assign.
    """
    import pandas as pd

    ZONE_LABELS = {'contact fields': 'contact', 'team': 'team',
                   'app': 'app', 'projects': 'project', 'tags': 'tag'}
    TRUTHY = {'y', 'yes', '1', 'true', 'x', 'v'}

    sname = next((s for s in xl.sheet_names if 'Full Info' in s or 'Full_Info' in s), None)
    if not sname:
        return

    df_raw = xl.parse(sname, header=None).fillna('')
    if len(df_raw) < 3:
        return

    row0 = [str(v).strip() for v in df_raw.iloc[0]]  # zone labels
    row1 = [str(v).strip() for v in df_raw.iloc[1]]  # column names

    # Build zone map: col_idx → zone key
    zone_map, cur = [], 'contact'
    for v in row0:
        lv = v.lower()
        if lv in ZONE_LABELS:
            cur = ZONE_LABELS[lv]
        zone_map.append(cur)

    # Build column lookup: zone → {col_name: col_idx}
    zone_cols = {'contact': {}, 'app': {}, 'project': {}, 'tag': {}}
    for ci, (zone, name) in enumerate(zip(zone_map, row1)):
        if zone in zone_cols and name:
            zone_cols[zone][name] = ci

    # Auto-create new master entities from new column headers
    palette = ['#3B82F6','#10B981','#F59E0B','#EF4444','#8B5CF6','#F97316','#14B8A6','#0EA5E9']
    existing_apps    = {r['name'] for r in conn.execute("SELECT name FROM master_apps").fetchall()}
    existing_projs   = {r['name'] for r in conn.execute("SELECT name FROM master_projects").fetchall()}
    existing_tags    = {r['name'] for r in conn.execute("SELECT name FROM tags").fetchall()}

    new_count = {'apps':0,'projects':0,'tags':0}
    for name in zone_cols['app']:
        if name not in existing_apps:
            c = palette[new_count['apps'] % len(palette)]
            conn.execute("INSERT OR IGNORE INTO master_apps(name,color) VALUES(?,?)", (name, c))
            new_count['apps'] += 1
    for name in zone_cols['project']:
        if name not in existing_projs:
            c = palette[new_count['projects'] % len(palette)]
            conn.execute("INSERT OR IGNORE INTO master_projects(name,color) VALUES(?,?)", (name, c))
            new_count['projects'] += 1
    for name in zone_cols['tag']:
        if name not in existing_tags:
            conn.execute("INSERT OR IGNORE INTO tags(name,color) VALUES(?,?)", (name, '#64748B'))
            new_count['tags'] += 1
    conn.commit()

    id_col = zone_cols['contact'].get('ContactID', 0)

    for _, raw_row in df_raw.iloc[2:].iterrows():
        vals = [str(v).strip() for v in raw_row]
        if not any(vals):
            continue

        cid     = vals[id_col] if id_col < len(vals) else ''
        name_th = vals[zone_cols['contact'].get('ชื่อ (ไทย)', -1)] if zone_cols['contact'].get('ชื่อ (ไทย)') is not None else ''
        if not name_th:
            continue

        # Generate ID if blank
        if not cid or cid in ('nan', ''):
            cid = _next_id(conn)
            result.setdefault('new_ids_assigned', []).append(cid)

        def cv(col_name):
            ci = zone_cols['contact'].get(col_name)
            return vals[ci] if ci is not None and ci < len(vals) else ''

        excel_dt = _parse_dt(cv('Updated'))
        db_row   = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()

        if not db_row:
            conn.execute("""INSERT INTO contacts
                (id,name_th,name_en,nickname,team,sub_team,org_role,direct_report,
                 email1,email2,phone,line_id,note_short,general_note,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cid, name_th, cv('Name (Eng)'), cv('ชื่อเล่น'), cv('Team'), cv('Sub-Team'),
                 cv('ตำแหน่ง'), cv('Direct Report'), cv('Email (หลัก)'), cv('Email (2)'),
                 cv('โทรศัพท์'), cv('Line ID'), cv('หมายเหตุสั้น'), cv('General Note'),
                 now_str, now_str))
            result['new_in_excel'] += 1
        else:
            db_dt = _parse_dt(str(db_row['updated_at']))
            excel_wins = True
            if excel_dt and db_dt:
                diff = (db_dt - excel_dt).total_seconds()
                if diff > bias_s:
                    excel_wins = False
                    result['conflicts'].append({'id': cid, 'name': name_th,
                        'winner': 'db', 'excel_updated': str(excel_dt)[:16],
                        'db_updated': str(db_dt)[:16]})
            if excel_wins:
                conn.execute("""UPDATE contacts SET
                    name_th=?,name_en=?,nickname=?,team=?,sub_team=?,org_role=?,direct_report=?,
                    email1=?,email2=?,phone=?,line_id=?,note_short=?,general_note=?,updated_at=?
                    WHERE id=?""",
                    (name_th, cv('Name (Eng)'), cv('ชื่อเล่น'), cv('Team'), cv('Sub-Team'),
                     cv('ตำแหน่ง'), cv('Direct Report'), cv('Email (หลัก)'), cv('Email (2)'),
                     cv('โทรศัพท์'), cv('Line ID'), cv('หมายเหตุสั้น'), cv('General Note'),
                     now_str, cid))
                result['from_excel'] += 1
            else:
                result['to_excel'] += 1

        # Assignments — Excel is master (replace DB state)
        def is_assigned(v):
            return v.lower() in TRUTHY or (v and v.lower() not in ('', 'nan', 'none', '0', 'false'))

        # Apps (value = role)
        conn.execute("DELETE FROM contact_apps WHERE contact_id=?", (cid,))
        for name, ci in zone_cols['app'].items():
            if ci < len(vals):
                role_val = vals[ci].strip()
                if is_assigned(role_val):
                    role = role_val if role_val.lower() not in TRUTHY else 'Y'
                    conn.execute("INSERT OR IGNORE INTO contact_apps(contact_id,app_name,role) VALUES(?,?,?)",
                                 (cid, name, role))

        # Projects (value = role)
        conn.execute("DELETE FROM contact_projects WHERE contact_id=?", (cid,))
        for name, ci in zone_cols['project'].items():
            if ci < len(vals):
                role_val = vals[ci].strip()
                if is_assigned(role_val):
                    role = role_val if role_val.lower() not in TRUTHY else 'Member'
                    conn.execute("INSERT INTO contact_projects(contact_id,project_name,role) VALUES(?,?,?)",
                                 (cid, name, role))
                    result['projects_synced'] += 1

        # Tags
        conn.execute("DELETE FROM contact_tags WHERE contact_id=?", (cid,))
        for name, ci in zone_cols['tag'].items():
            if ci < len(vals) and is_assigned(vals[ci]):
                tag_row = conn.execute("SELECT id FROM tags WHERE name=?", (name,)).fetchone()
                if tag_row:
                    conn.execute("INSERT OR IGNORE INTO contact_tags VALUES(?,?)", (cid, tag_row['id']))

    conn.commit()
    if any(new_count.values()):
        result.setdefault('new_entities', new_count)

def _build_workbook(conn):
    """Build and return an openpyxl Workbook with all export sheets."""
    from openpyxl import Workbook
    wb = Workbook()
    ws_fi = wb.active
    ws_fi.title = '📊 Full Info'
    _build_fullinfo_sheet(ws_fi, conn)
    _build_notes_sheet(wb.create_sheet('📝 Notes'), conn)
    _build_master_sheet(wb.create_sheet('⚙ Master Data'), conn)
    return wb

def _write_excel_to_path(path, conn):
    """Export full DB to Excel (3 sheets). Returns actual path written."""
    wb = _build_workbook(conn)
    target = str(path)
    try:
        wb.save(target)
        return target
    except PermissionError:
        from pathlib import Path
        p = Path(path)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fallback = str(p.parent / f"{p.stem}_sync_{ts}{p.suffix}")
        wb.save(fallback)
        return fallback

def _do_twoway_sync(excel_path, conn, write_back=True):
    """
    Two-way sync: Full Info sheet (primary) + Notes (append-only).
    Falls back to legacy Contacts sheet if Full Info not found.
    """
    import pandas as pd

    BIAS_SECONDS = 60
    result = {
        'from_excel': 0, 'to_excel': 0,
        'new_in_excel': 0, 'new_in_db': 0,
        'notes_appended': 0, 'projects_synced': 0,
        'conflicts': [], 'errors': [],
    }

    try:
        xl = pd.ExcelFile(excel_path)
    except Exception as e:
        result['errors'].append(f'อ่าน Excel ไม่ได้: {e}')
        return result

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── 1. Primary: Full Info sheet ───────────────────────────────────────────
    has_fullinfo = any('Full Info' in s or 'Full_Info' in s for s in xl.sheet_names)
    if has_fullinfo:
        try:
            _parse_fullinfo_sheet(xl, conn, result, now_str, BIAS_SECONDS)
        except Exception as e:
            result['errors'].append(f'Full Info sync error: {e}')
    else:
        # ── Legacy fallback: Contacts sheet ───────────────────────────────────
        sname = next((s for s in xl.sheet_names if 'Contacts' in s), None)
        if sname:
            df = xl.parse(sname).fillna('')
            df = df.rename(columns=CONTACT_COL_MAP)
            excel_ids = set()
            for _, row in df.iterrows():
                cid = _str(row, 'id'); name_th = _str(row, 'name_th')
                if not cid or not name_th: continue
                excel_ids.add(cid)
                excel_dt = _parse_dt(_str(row, 'excel_updated'))
                db_row   = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()
                if not db_row:
                    conn.execute("""INSERT INTO contacts
                        (id,name_th,name_en,nickname,team,org_role,direct_report,
                         email1,email2,phone,line_id,note_short,general_note,created_at,updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (cid, name_th,
                         _str(row,'name_en'), _str(row,'nickname'), _str(row,'team'),
                         _str(row,'org_role'), _str(row,'direct_report'), _str(row,'email1'),
                         _str(row,'email2'), _str(row,'phone'), _str(row,'line_id'),
                         _str(row,'note_short'), _str(row,'general_note'), now_str, now_str))
                    result['new_in_excel'] += 1
                else:
                    db_dt = _parse_dt(str(db_row['updated_at']))
                    excel_wins = True
                    if excel_dt and db_dt:
                        diff = (db_dt - excel_dt).total_seconds()
                        if diff > BIAS_SECONDS:
                            excel_wins = False
                            result['conflicts'].append({'id': cid, 'name': name_th, 'winner': 'db',
                                'excel_updated': str(excel_dt)[:16], 'db_updated': str(db_dt)[:16]})
                    if excel_wins:
                        conn.execute("""UPDATE contacts SET
                            name_th=?,name_en=?,nickname=?,team=?,org_role=?,direct_report=?,
                            email1=?,email2=?,phone=?,line_id=?,note_short=?,general_note=?,updated_at=?
                            WHERE id=?""",
                            (name_th, _str(row,'name_en'), _str(row,'nickname'), _str(row,'team'),
                             _str(row,'org_role'), _str(row,'direct_report'), _str(row,'email1'),
                             _str(row,'email2'), _str(row,'phone'), _str(row,'line_id'),
                             _str(row,'note_short'), _str(row,'general_note'), now_str, cid))
                        result['from_excel'] += 1
                    else:
                        result['to_excel'] += 1
            all_db_ids = {r['id'] for r in conn.execute("SELECT id FROM contacts").fetchall()}
            result['new_in_db'] = len(all_db_ids - excel_ids)
            conn.commit()

    # ── 2. Notes (append-only, both paths) ───────────────────────────────────
    sname_n = next((s for s in xl.sheet_names if 'Note' in s), None)
    if sname_n:
        df_n = xl.parse(sname_n, header=0).fillna('')
        df_n = df_n.rename(columns={'ContactID':'cid','วันที่':'note_date',
                                     'หัวข้อ':'title','เนื้อหา Note':'content'})
        for _, row in df_n.iterrows():
            cid = _str(row, 'cid'); content = _str(row, 'content')
            if not cid or not content: continue
            if not conn.execute("SELECT id FROM contacts WHERE id=?", (cid,)).fetchone(): continue
            if not conn.execute("SELECT id FROM notes WHERE contact_id=? AND content=?",
                                (cid, content)).fetchone():
                conn.execute("INSERT INTO notes (contact_id,note_date,title,content) VALUES (?,?,?,?)",
                             (cid, _str(row,'note_date'), _str(row,'title'), content))
                result['notes_appended'] += 1
        conn.commit()

    # ── 3. Count DB-only contacts (for stats) ────────────────────────────────
    if not has_fullinfo:
        pass  # already counted in legacy path
    else:
        result['new_in_db'] = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]

    # ── 4. Write merged result back to Excel ──────────────────────────────────
    if write_back:
        try:
            written_path = _write_excel_to_path(excel_path, conn)
            result['excel_written'] = True
            if written_path != str(excel_path):
                result['excel_locked'] = True
                result['fallback_path'] = written_path
        except Exception as e:
            result['excel_written'] = False
            result['errors'].append(f'เขียน Excel ไม่ได้: {e}')
    else:
        result['excel_written'] = False

    return result

# ─── Main page ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('app.html')

# ─── Contacts ──────────────────────────────────────────────────────────────────

@app.route('/api/contacts', methods=['GET'])
def get_contacts():
    q    = request.args.get('q', '').strip()
    team = request.args.get('team', '').strip()
    proj = request.args.get('project', '').strip()
    tag  = request.args.get('tag', '').strip()

    conn = db_conn()
    sql  = "SELECT DISTINCT c.* FROM contacts c"
    params, where = [], []

    if proj:
        sql += " JOIN contact_projects cp ON cp.contact_id=c.id"
        where.append("cp.project_name=?"); params.append(proj)
    if tag:
        sql += " JOIN contact_tags ct ON ct.contact_id=c.id JOIN tags t ON t.id=ct.tag_id"
        where.append("t.name=?"); params.append(tag)
    if team:
        where.append("c.team=?"); params.append(team)
    if q:
        where.append("(c.name_th LIKE ? OR c.name_en LIKE ? OR c.nickname LIKE ? OR c.email1 LIKE ? OR c.note_short LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like, like, like]

    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY c.name_th"

    rows = conn.execute(sql, params).fetchall()
    data = [_contact_dict(r, conn) for r in rows]
    conn.close()
    return jsonify(data)

@app.route('/api/contacts', methods=['POST'])
def create_contact():
    body = request.get_json(force=True)
    conn = db_conn()
    cid  = _next_id(conn)
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("""INSERT INTO contacts
        (id,name_th,name_en,nickname,team,sub_team,org_role,direct_report,
         email1,email2,phone,line_id,note_short,general_note,work_note,quick_note_team,
         created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
        cid,
        body.get('name_th',''),       body.get('name_en'),         body.get('nickname'),
        body.get('team'),             body.get('sub_team'),        body.get('org_role'),
        body.get('direct_report'),    body.get('email1'),          body.get('email2'),
        body.get('phone'),            body.get('line_id'),         body.get('note_short'),
        body.get('general_note'),     body.get('work_note'),       body.get('quick_note_team'),
        now, now,
    ))
    for p in body.get('projects', []):
        conn.execute("INSERT INTO contact_projects (contact_id,project_name,role,note,sort_order) VALUES (?,?,?,?,?)",
                     (cid, p['project_name'], p.get('role','Supporter'), p.get('note'), p.get('sort_order',0)))
    for tid in body.get('tag_ids', []):
        conn.execute("INSERT OR IGNORE INTO contact_tags VALUES (?,?)", (cid, tid))
    for rname in body.get('roles', []):
        conn.execute("INSERT OR IGNORE INTO contact_roles VALUES (?,?)", (cid, rname))
    entity = body.get('entity')
    if entity and entity.get('entity_type'):
        conn.execute("INSERT OR REPLACE INTO contact_entity VALUES (?,?,?)",
                     (cid, entity['entity_type'], entity.get('entity_value','')))
    for sname in body.get('systems', []):
        conn.execute("INSERT OR IGNORE INTO contact_systems VALUES (?,?)", (cid, sname))
    conn.commit()
    row  = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()
    data = _contact_dict(row, conn, full=True)
    conn.close()
    return jsonify(data), 201

@app.route('/api/contacts/<cid>', methods=['GET'])
def get_contact(cid):
    conn = db_conn()
    row  = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    data = _contact_dict(row, conn, full=True)
    conn.close()
    return jsonify(data)

@app.route('/api/contacts/<cid>', methods=['PUT'])
def update_contact(cid):
    body = request.get_json(force=True)
    conn = db_conn()
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    fields = ['name_th','name_en','nickname','team','sub_team','org_role','direct_report',
              'email1','email2','phone','line_id','note_short','general_note',
              'work_note','quick_note_team','associated_to_meeting']
    sets   = [f"{f}=?" for f in fields] + ['updated_at=?']
    vals   = [body.get(f) for f in fields] + [now, cid]
    conn.execute(f"UPDATE contacts SET {','.join(sets)} WHERE id=?", vals)

    if 'projects' in body:
        conn.execute("DELETE FROM contact_projects WHERE contact_id=?", (cid,))
        conn.execute("DELETE FROM contact_project_subvalues WHERE contact_id=?", (cid,))
        for i, p in enumerate(body['projects']):
            conn.execute("INSERT INTO contact_projects (contact_id,project_name,role,note,sort_order) VALUES (?,?,?,?,?)",
                         (cid, p['project_name'], p.get('role','Supporter'), p.get('note'), i))
            for col_name, value in (p.get('subvalues') or {}).items():
                if value and str(value).strip():
                    conn.execute("INSERT OR REPLACE INTO contact_project_subvalues VALUES (?,?,?,?)",
                                 (cid, p['project_name'], col_name, str(value).strip()))

    if 'tag_ids' in body:
        conn.execute("DELETE FROM contact_tags WHERE contact_id=?", (cid,))
        for tid in body['tag_ids']:
            conn.execute("INSERT OR IGNORE INTO contact_tags VALUES (?,?)", (cid, tid))

    if 'roles' in body:
        conn.execute("DELETE FROM contact_roles WHERE contact_id=?", (cid,))
        for rname in body['roles']:
            conn.execute("INSERT OR IGNORE INTO contact_roles VALUES (?,?)", (cid, rname))

    if 'entity' in body:
        conn.execute("DELETE FROM contact_entity WHERE contact_id=?", (cid,))
        ent = body['entity']
        if ent and ent.get('entity_type'):
            conn.execute("INSERT INTO contact_entity VALUES (?,?,?)",
                         (cid, ent['entity_type'], ent.get('entity_value','')))

    if 'systems' in body:
        conn.execute("DELETE FROM contact_systems WHERE contact_id=?", (cid,))
        for sname in body['systems']:
            conn.execute("INSERT OR IGNORE INTO contact_systems VALUES (?,?)", (cid, sname))

    conn.commit()
    row  = conn.execute("SELECT * FROM contacts WHERE id=?", (cid,)).fetchone()
    data = _contact_dict(row, conn, full=True)
    conn.close()
    return jsonify(data)

@app.route('/api/contacts/<cid>', methods=['DELETE'])
def delete_contact(cid):
    conn = db_conn()
    conn.execute("DELETE FROM contacts WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    # clean up photo file
    _fp = os.path.join(AVATARS_DIR, f'{cid}.jpg')
    if os.path.exists(_fp):
        os.remove(_fp)
    return jsonify({'ok': True})

# ─── Photo upload / delete ─────────────────────────────────────────────────────

@app.route('/api/contacts/<cid>/photo', methods=['POST'])
def upload_photo(cid):
    if 'photo' not in request.files:
        return jsonify({'error': 'no photo field'}), 400
    f    = request.files['photo']
    dest = os.path.join(AVATARS_DIR, f'{cid}.jpg')
    try:
        if _PIL_OK:
            img = _PILImage.open(f.stream).convert('RGB')
            img = img.resize((400, 400), _PILImage.LANCZOS)
            img.save(dest, 'JPEG', quality=88, optimize=True)
        else:
            f.save(dest)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    conn = db_conn()
    conn.execute("UPDATE contacts SET photo_path=?, updated_at=? WHERE id=?",
                 (f'{cid}.jpg', datetime.now().isoformat(), cid))
    conn.commit(); conn.close()
    ts = int(os.path.getmtime(dest))
    return jsonify({'ok': True, 'photo_url': f'/static/avatars/{cid}.jpg?v={ts}'})

@app.route('/api/contacts/<cid>/photo', methods=['DELETE'])
def delete_photo(cid):
    conn = db_conn()
    conn.execute("UPDATE contacts SET photo_path=NULL WHERE id=?", (cid,))
    conn.commit(); conn.close()
    _fp = os.path.join(AVATARS_DIR, f'{cid}.jpg')
    if os.path.exists(_fp):
        os.remove(_fp)
    return jsonify({'ok': True})

# ─── Notes ─────────────────────────────────────────────────────────────────────

@app.route('/api/contacts/<cid>/notes', methods=['POST'])
def add_note(cid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("INSERT INTO notes (contact_id,note_date,title,content) VALUES (?,?,?,?)",
                 (cid, body.get('note_date', datetime.now().strftime('%Y-%m-%d')),
                  body.get('title'), body['content']))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = conn.execute("SELECT * FROM notes WHERE id=?", (nid,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201

@app.route('/api/notes/<int:nid>', methods=['PUT'])
def update_note(nid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("UPDATE notes SET note_date=?,title=?,content=? WHERE id=?",
                 (body.get('note_date'), body.get('title'), body['content'], nid))
    conn.commit()
    row = conn.execute("SELECT * FROM notes WHERE id=?", (nid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/notes/<int:nid>', methods=['DELETE'])
def delete_note(nid):
    conn = db_conn()
    conn.execute("DELETE FROM notes WHERE id=?", (nid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Todos ─────────────────────────────────────────────────────────────────────

@app.route('/api/contacts/<cid>/todos', methods=['GET'])
def get_todos(cid):
    conn = db_conn()
    rows = conn.execute("SELECT * FROM todos WHERE contact_id=? ORDER BY done, due_date, id", (cid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/contacts/<cid>/todos', methods=['POST'])
def create_todo(cid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("INSERT INTO todos (contact_id, title, due_date) VALUES (?,?,?)",
                 (cid, body['title'], body.get('due_date') or None))
    conn.commit()
    row = conn.execute("SELECT * FROM todos WHERE id=last_insert_rowid()").fetchone()
    conn.close()
    return jsonify(dict(row)), 201

@app.route('/api/todos/<int:tid>', methods=['PUT'])
def update_todo(tid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("UPDATE todos SET title=?, done=?, due_date=? WHERE id=?",
                 (body.get('title'), 1 if body.get('done') else 0,
                  body.get('due_date') or None, tid))
    conn.commit()
    row = conn.execute("SELECT * FROM todos WHERE id=?", (tid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/todos/<int:tid>', methods=['DELETE'])
def delete_todo(tid):
    conn = db_conn()
    conn.execute("DELETE FROM todos WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Tags ──────────────────────────────────────────────────────────────────────

@app.route('/api/tags', methods=['GET'])
def get_tags():
    conn = db_conn()
    rows = conn.execute("SELECT * FROM tags ORDER BY sort_order, name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/tags/reorder', methods=['PUT'])
def reorder_tags():
    order = request.get_json(force=True).get('order', [])
    conn  = db_conn()
    for i, tid in enumerate(order):
        conn.execute("UPDATE tags SET sort_order=? WHERE id=?", (i, int(tid)))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/tags', methods=['POST'])
def create_tag():
    body  = request.get_json(force=True)
    conn  = db_conn()
    palette = ['#8B5CF6','#F97316','#14B8A6','#0EA5E9','#F43F5E','#EC4899','#84CC16','#3B82F6']
    count = conn.execute("SELECT COUNT(*) FROM tags").fetchone()[0]
    color = body.get('color', palette[count % len(palette)])
    conn.execute("INSERT OR IGNORE INTO tags (name,color) VALUES (?,?)", (body['name'], color))
    conn.commit()
    row = conn.execute("SELECT * FROM tags WHERE name=?", (body['name'],)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201

# ─── Master Roles ──────────────────────────────────────────────────────────────

@app.route('/api/master/roles', methods=['GET'])
def get_master_roles():
    conn = db_conn()
    rows = conn.execute("SELECT * FROM master_roles ORDER BY sort_order, name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/master/roles/reorder', methods=['PUT'])
def reorder_roles():
    order = request.get_json(force=True).get('order', [])
    conn  = db_conn()
    for i, name in enumerate(order):
        conn.execute("UPDATE master_roles SET sort_order=? WHERE name=?", (i, name))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/roles', methods=['POST'])
def create_master_role():
    body = request.get_json(force=True)
    conn = db_conn()
    max_ord = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM master_roles").fetchone()[0]
    conn.execute("INSERT OR IGNORE INTO master_roles (name,color,sort_order) VALUES (?,?,?)",
                 (body['name'], body.get('color','#64748B'), max_ord + 1))
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/master/roles/<name>', methods=['PUT'])
def update_master_role(name):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("UPDATE master_roles SET name=?,color=?,sort_order=? WHERE name=?",
                 (body.get('name',name), body.get('color','#64748B'), body.get('sort_order',0), name))
    if body.get('name') and body['name'] != name:
        conn.execute("UPDATE contact_roles SET role_name=? WHERE role_name=?", (body['name'], name))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/roles/<name>', methods=['DELETE'])
def delete_master_role(name):
    conn = db_conn()
    conn.execute("DELETE FROM contact_roles WHERE role_name=?", (name,))
    conn.execute("DELETE FROM master_roles WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Contact Roles ─────────────────────────────────────────────────────────────

@app.route('/api/contacts/<cid>/roles', methods=['PUT'])
def set_contact_roles(cid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("DELETE FROM contact_roles WHERE contact_id=?", (cid,))
    for rname in body.get('roles', []):
        conn.execute("INSERT OR IGNORE INTO contact_roles VALUES (?,?)", (cid, rname))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Contact Entity ─────────────────────────────────────────────────────────────

ENTITY_TYPES = ['KS', 'Vendor', 'Subsidiary', '3rd Party', 'Other Bank']

@app.route('/api/contacts/<cid>/entity', methods=['PUT'])
def set_contact_entity(cid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("DELETE FROM contact_entity WHERE contact_id=?", (cid,))
    if body.get('entity_type'):
        conn.execute("INSERT INTO contact_entity VALUES (?,?,?)",
                     (cid, body['entity_type'], body.get('entity_value','')))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Master Systems ─────────────────────────────────────────────────────────────

@app.route('/api/master/systems', methods=['GET'])
def get_master_systems():
    conn = db_conn()
    rows = conn.execute("SELECT * FROM master_systems ORDER BY sort_order, name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/master/systems/reorder', methods=['PUT'])
def reorder_systems():
    order = request.get_json(force=True).get('order', [])
    conn  = db_conn()
    for i, name in enumerate(order):
        conn.execute("UPDATE master_systems SET sort_order=? WHERE name=?", (i, name))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/systems', methods=['POST'])
def create_master_system():
    body = request.get_json(force=True)
    conn = db_conn()
    max_ord = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM master_systems").fetchone()[0]
    conn.execute("INSERT OR IGNORE INTO master_systems (name,sort_order) VALUES (?,?)",
                 (body['name'], max_ord + 1))
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/master/systems/<name>', methods=['DELETE'])
def delete_master_system(name):
    conn = db_conn()
    conn.execute("DELETE FROM contact_systems WHERE system_name=?", (name,))
    conn.execute("DELETE FROM master_systems WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Contact Systems ────────────────────────────────────────────────────────────

@app.route('/api/contacts/<cid>/systems', methods=['PUT'])
def set_contact_systems(cid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("DELETE FROM contact_systems WHERE contact_id=?", (cid,))
    for sname in body.get('systems', []):
        conn.execute("INSERT OR IGNORE INTO contact_systems VALUES (?,?)", (cid, sname))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Project Sub-columns ────────────────────────────────────────────────────────

@app.route('/api/projects/<proj>/subcolumns', methods=['GET'])
def get_project_subcolumns(proj):
    conn = db_conn()
    rows = conn.execute(
        "SELECT * FROM project_subcolumns WHERE project_name=? ORDER BY sort_order, id",
        (proj,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/projects/<proj>/subcolumns', methods=['POST'])
def add_project_subcolumn(proj):
    body = request.get_json(force=True)
    conn = db_conn()
    max_ord = conn.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM project_subcolumns WHERE project_name=?",
        (proj,)).fetchone()[0]
    conn.execute("INSERT OR IGNORE INTO project_subcolumns (project_name,col_name,sort_order,visible) VALUES (?,?,?,1)",
                 (proj, body['col_name'], max_ord + 1))
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/projects/<proj>/subcolumns/<int:scid>', methods=['PUT'])
def update_project_subcolumn(proj, scid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("UPDATE project_subcolumns SET col_name=?,sort_order=?,visible=? WHERE id=? AND project_name=?",
                 (body.get('col_name'), body.get('sort_order',0), body.get('visible',1), scid, proj))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/projects/<proj>/subcolumns/<int:scid>', methods=['DELETE'])
def delete_project_subcolumn(proj, scid):
    conn = db_conn()
    row = conn.execute("SELECT col_name FROM project_subcolumns WHERE id=? AND project_name=?", (scid, proj)).fetchone()
    if row:
        conn.execute("DELETE FROM contact_project_subvalues WHERE project_name=? AND col_name=?",
                     (proj, row['col_name']))
    conn.execute("DELETE FROM project_subcolumns WHERE id=? AND project_name=?", (scid, proj))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Contact Project Sub-values ─────────────────────────────────────────────────

@app.route('/api/contacts/<cid>/projects/<proj>/subvalues', methods=['PUT'])
def set_project_subvalues(cid, proj):
    body = request.get_json(force=True)
    conn = db_conn()
    for col_name, value in body.get('subvalues', {}).items():
        if value:
            conn.execute("INSERT OR REPLACE INTO contact_project_subvalues VALUES (?,?,?,?)",
                         (cid, proj, col_name, value))
        else:
            conn.execute("DELETE FROM contact_project_subvalues WHERE contact_id=? AND project_name=? AND col_name=?",
                         (cid, proj, col_name))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Master data ───────────────────────────────────────────────────────────────

@app.route('/api/master/projects')
def master_projects():
    conn = db_conn()
    rows = conn.execute("SELECT * FROM master_projects ORDER BY sort_order, name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/master/projects/reorder', methods=['PUT'])
def reorder_projects():
    order = request.get_json(force=True).get('order', [])
    conn  = db_conn()
    for i, name in enumerate(order):
        conn.execute("UPDATE master_projects SET sort_order=? WHERE name=?", (i, name))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/projects', methods=['POST'])
def add_project():
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("INSERT OR REPLACE INTO master_projects (name,color,short_name) VALUES (?,?,?)",
                 (body['name'], body.get('color','#64748B'), body.get('short_name')))
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

@app.route('/api/master/projects/<name>', methods=['PUT'])
def update_project(name):
    body = request.get_json(force=True)
    new_name  = body.get('name', name)
    color     = body.get('color', '#64748B')
    short_name= body.get('short_name', '')
    conn = db_conn()
    conn.execute("UPDATE master_projects SET name=?,color=?,short_name=? WHERE name=?",
                 (new_name, color, short_name, name))
    if new_name != name:
        conn.execute("UPDATE contact_projects SET project_name=? WHERE project_name=?", (new_name, name))
        conn.execute("UPDATE project_subcolumns SET project_name=? WHERE project_name=?", (new_name, name))
        conn.execute("UPDATE contact_project_subvalues SET project_name=? WHERE project_name=?", (new_name, name))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/projects/<name>', methods=['DELETE'])
def delete_project(name):
    conn = db_conn()
    conn.execute("DELETE FROM contact_project_subvalues WHERE project_name=?", (name,))
    conn.execute("DELETE FROM project_subcolumns WHERE project_name=?", (name,))
    conn.execute("DELETE FROM contact_projects WHERE project_name=?", (name,))
    conn.execute("DELETE FROM master_projects WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/systems/<name>', methods=['PUT'])
def update_master_system(name):
    body = request.get_json(force=True)
    new_name = body.get('name', name)
    conn = db_conn()
    conn.execute("UPDATE master_systems SET name=? WHERE name=?", (new_name, name))
    if new_name != name:
        conn.execute("UPDATE contact_systems SET system_name=? WHERE system_name=?", (new_name, name))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/tags/<int:tid>', methods=['PUT'])
def update_tag(tid):
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("UPDATE tags SET name=?,color=? WHERE id=?",
                 (body.get('name'), body.get('color','#64748B'), tid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/tags/<int:tid>', methods=['DELETE'])
def delete_tag(tid):
    conn = db_conn()
    conn.execute("DELETE FROM contact_tags WHERE tag_id=?", (tid,))
    conn.execute("DELETE FROM tags WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/master/teams')
def master_teams():
    conn = db_conn()
    rows = conn.execute("SELECT * FROM master_teams ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/master/apps')
def master_apps():
    conn = db_conn()
    rows = conn.execute("SELECT * FROM master_apps ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/master/apps', methods=['POST'])
def add_app():
    body = request.get_json(force=True)
    conn = db_conn()
    conn.execute("INSERT OR REPLACE INTO master_apps (name,color) VALUES (?,?)",
                 (body['name'], body.get('color', '#6366F1')))
    conn.commit()
    conn.close()
    return jsonify({'ok': True}), 201

# ─── Excel Sync ────────────────────────────────────────────────────────────────

@app.route('/api/sync/export', methods=['GET'])
def export_excel():
    """Download a snapshot of the full DB as .xlsx with Full Info sheet."""
    conn = db_conn()
    wb = _build_workbook(conn)
    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mycontacts_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/sync/export_view', methods=['POST'])
def export_excel_view():
    """Export custom-column Excel from grid view (headers + rows sent by client)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    body    = request.get_json(force=True)
    headers = body.get('headers', [])
    rows    = body.get('rows', [])
    wb = Workbook()
    ws = wb.active
    ws.title = 'View'
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type='solid', fgColor='D1D5DB')
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mycontacts_view_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── Settings ──────────────────────────────────────────────────────────────────

@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = db_conn()
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    conn.close()
    return jsonify({r['key']: r['value'] for r in rows})

@app.route('/api/settings', methods=['POST'])
def save_settings():
    body = request.get_json(force=True)
    conn = db_conn()
    for key, value in body.items():
        conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (key, str(value)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ─── Two-Way Sync ──────────────────────────────────────────────────────────────

@app.route('/api/sync/sync', methods=['POST'])
def two_way_sync():
    """Two-way sync: merge DB ↔ Excel file, Excel-biased conflict resolution."""
    body = request.get_json(force=True, silent=True) or {}
    conn = db_conn()

    # Resolve Excel path: from request body first, then saved setting
    excel_path = body.get('excel_path', '').strip()
    if not excel_path:
        row = conn.execute("SELECT value FROM settings WHERE key='excel_path'").fetchone()
        excel_path = row['value'] if row else ''

    if not excel_path:
        conn.close()
        return jsonify({'error': 'ยังไม่ได้ตั้ง Excel Path — กรุณาตั้งค่าก่อน'}), 400

    if not os.path.exists(excel_path):
        conn.close()
        return jsonify({'error': f'ไม่พบไฟล์: {excel_path}'}), 400

    # Save path if provided in request
    if body.get('excel_path'):
        conn.execute("INSERT OR REPLACE INTO settings VALUES ('excel_path',?)", (excel_path,))
        conn.commit()

    result = _do_twoway_sync(excel_path, conn)

    # Record sync log + last_sync timestamp
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("""INSERT INTO sync_log
        (synced_at, excel_path, from_excel, to_excel, new_in_excel, new_in_db,
         notes_appended, projects_synced, conflicts, errors)
        VALUES (?,?,?,?,?,?,?,?,?,?)""", (
        now_str, excel_path,
        result['from_excel'], result['to_excel'],
        result['new_in_excel'], result['new_in_db'],
        result['notes_appended'], result['projects_synced'],
        json.dumps(result['conflicts'], ensure_ascii=False),
        json.dumps(result['errors'],    ensure_ascii=False),
    ))
    conn.execute("INSERT OR REPLACE INTO settings VALUES ('last_sync',?)", (now_str,))
    conn.execute("INSERT OR REPLACE INTO settings VALUES ('last_sync_path',?)", (excel_path,))
    conn.commit()
    conn.close()

    return jsonify({'ok': True, 'synced_at': now_str, **result})

@app.route('/api/sync/last', methods=['GET'])
def sync_last():
    """Return last sync info + recent sync_log entries."""
    conn = db_conn()
    rows = conn.execute("SELECT key,value FROM settings WHERE key IN ('last_sync','excel_path','last_sync_path')").fetchall()
    settings = {r['key']: r['value'] for r in rows}
    logs = [dict(r) for r in conn.execute(
        "SELECT * FROM sync_log ORDER BY id DESC LIMIT 10").fetchall()]
    conn.close()
    return jsonify({'settings': settings, 'logs': logs})

@app.route('/api/sync/import', methods=['POST'])
def import_excel():
    """One-way import from uploaded .xlsx — supports Full Info sheet and legacy format."""
    import pandas as pd

    if 'file' not in request.files:
        return jsonify({'error': 'ไม่พบไฟล์ — กรุณา upload .xlsx'}), 400
    try:
        xl = pd.ExcelFile(request.files['file'])
    except Exception as e:
        return jsonify({'error': f'อ่านไฟล์ไม่ได้: {e}'}), 400

    conn = db_conn()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result  = {'from_excel': 0, 'to_excel': 0, 'new_in_excel': 0, 'new_in_db': 0,
               'notes_appended': 0, 'projects_synced': 0, 'conflicts': [], 'errors': []}

    has_fullinfo = any('Full Info' in s or 'Full_Info' in s for s in xl.sheet_names)
    if has_fullinfo:
        _parse_fullinfo_sheet(xl, conn, result, now_str, bias_s=0)  # bias=0 → Excel always wins
    else:
        # Legacy Contacts sheet
        sname = next((s for s in xl.sheet_names if 'Contacts' in s), None)
        if sname:
            df = xl.parse(sname).fillna('').rename(columns={
                'ContactID':'id','ชื่อ (ไทย)':'name_th','Name (Eng)':'name_en',
                'ชื่อเล่น':'nickname','ทีม':'team','ตำแหน่ง':'org_role',
                'Direct Report':'direct_report','Email (หลัก)':'email1','Email (2)':'email2',
                'โทรศัพท์':'phone','Line ID':'line_id','หมายเหตุสั้น':'note_short',
                'General Note':'general_note'})
            for _, row in df.iterrows():
                cid = str(row.get('id','')).strip()
                name_th = str(row.get('name_th','')).strip()
                if not cid or not name_th: continue
                vals = (name_th, str(row.get('name_en','')), str(row.get('nickname','')),
                        str(row.get('team','')), str(row.get('org_role','')), str(row.get('direct_report','')),
                        str(row.get('email1','')), str(row.get('email2','')), str(row.get('phone','')),
                        str(row.get('line_id','')), str(row.get('note_short','')), str(row.get('general_note','')))
                if conn.execute("SELECT id FROM contacts WHERE id=?", (cid,)).fetchone():
                    conn.execute("UPDATE contacts SET name_th=?,name_en=?,nickname=?,team=?,org_role=?,"
                                 "direct_report=?,email1=?,email2=?,phone=?,line_id=?,note_short=?,"
                                 "general_note=?,updated_at=? WHERE id=?", (*vals, now_str, cid))
                else:
                    conn.execute("INSERT INTO contacts (id,name_th,name_en,nickname,team,org_role,"
                                 "direct_report,email1,email2,phone,line_id,note_short,general_note,"
                                 "created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                 (cid, *vals, now_str, now_str))
                result['new_in_excel'] += 1
        conn.commit()

    imported = {'contacts': result['new_in_excel'] + result['from_excel'],
                'projects': result['projects_synced']}
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'imported': imported})

# ─── Stats ─────────────────────────────────────────────────────────────────────

@app.route('/api/stats')
def stats():
    conn = db_conn()
    data = {
        'total_contacts': conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0],
        'total_tags':     conn.execute("SELECT COUNT(*) FROM tags").fetchone()[0],
        'total_notes':    conn.execute("SELECT COUNT(*) FROM notes").fetchone()[0],
        'by_team': [dict(r) for r in conn.execute(
            "SELECT team, COUNT(*) as count FROM contacts GROUP BY team ORDER BY count DESC").fetchall()],
        'by_project': [dict(r) for r in conn.execute(
            "SELECT project_name, COUNT(*) as count FROM contact_projects GROUP BY project_name ORDER BY count DESC").fetchall()],
    }
    conn.close()
    return jsonify(data)

# ─── Auto File Watcher ─────────────────────────────────────────────────────────

class ExcelWatcher(threading.Thread):
    """
    Background thread: polls Excel file mtime every 3s.
    When user saves Excel → debounce 5s → auto two-way sync → notify browser.
    Ignores mtime changes caused by our own write-back (15s cooldown after sync).
    """
    POLL     = 3   # seconds between mtime checks
    DEBOUNCE = 5   # seconds of file-quiet before triggering sync
    COOLDOWN = 15  # seconds to ignore after we wrote Excel ourselves

    def __init__(self):
        super().__init__(daemon=True, name='ExcelWatcher')
        self._mtime      = None
        self._changed_at = 0
        self._skip_until = 0
        self.status      = 'idle'   # 'idle' | 'pending' | 'syncing'
        self.watching    = False
        self.last_sync_at = None
        self.last_result  = None

    def run(self):
        while True:
            try:
                self._tick()
            except Exception:
                self.status = 'idle'
            time.sleep(self.POLL)

    def _get_path(self):
        conn = db_conn()
        row  = conn.execute("SELECT value FROM settings WHERE key='excel_path'").fetchone()
        conn.close()
        return row['value'] if row else None

    def _tick(self):
        path = self._get_path()
        if not path or not os.path.exists(path):
            self.watching = False
            self.status   = 'idle'
            return

        self.watching = True
        mtime = os.path.getmtime(path)
        now   = time.time()

        # Init on first valid tick
        if self._mtime is None:
            self._mtime = mtime
            return

        # File changed and not in cooldown window (our own write-back)
        if mtime != self._mtime and now > self._skip_until:
            self._mtime      = mtime
            self._changed_at = now
            if self.status == 'idle':
                self.status = 'pending'

        # Debounce: file hasn't changed for DEBOUNCE seconds → sync
        if self.status == 'pending' and (now - self._changed_at) >= self.DEBOUNCE:
            self._do_sync(path)

    def _do_sync(self, path):
        self.status = 'syncing'
        conn   = db_conn()
        result = _do_twoway_sync(path, conn, write_back=False)
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn.execute("""INSERT INTO sync_log
            (synced_at, excel_path, from_excel, to_excel, new_in_excel, new_in_db,
             notes_appended, projects_synced, conflicts, errors)
            VALUES (?,?,?,?,?,?,?,?,?,?)""", (
            now_str, path,
            result['from_excel'], result['to_excel'],
            result['new_in_excel'], result['new_in_db'],
            result['notes_appended'], result['projects_synced'],
            json.dumps(result['conflicts'], ensure_ascii=False),
            json.dumps(result['errors'],    ensure_ascii=False),
        ))
        conn.execute("INSERT OR REPLACE INTO settings VALUES ('last_sync',?)", (now_str,))
        conn.commit()
        conn.close()
        # Refresh known mtime + set cooldown so write-back doesn't re-trigger
        if os.path.exists(path):
            self._mtime = os.path.getmtime(path)
        self._skip_until  = time.time() + self.COOLDOWN
        self.last_result  = result
        self.last_sync_at = now_str
        self.status       = 'idle'

watcher = ExcelWatcher()

@app.route('/api/sync/status')
def sync_status():
    return jsonify({
        'status':      watcher.status,
        'watching':    watcher.watching,
        'last_sync_at': watcher.last_sync_at,
        'last_result': watcher.last_result,
    })

# ─── Import Preview / Execute / Clear ──────────────────────────────────────────

def _parse_excel_for_preview(file_stream):
    """
    Read uploaded Excel Full Info sheet (3-row header) → return (rows, master_in_file, err).
    Parses all zones: INDIV, Projects (flag+sub-cols), Main Role, Entity, Systems, Tags.
    """
    import pandas as pd
    try:
        xl = pd.ExcelFile(file_stream)
    except Exception as e:
        return None, None, str(e)

    sname = next((s for s in xl.sheet_names if 'Full Info' in s or 'Full_Info' in s), None)
    if not sname and xl.sheet_names:
        sname = xl.sheet_names[0]
    if not sname:
        return None, None, 'ไม่พบ sheet ใน Excel'

    df = xl.parse(sname, header=None).fillna('')
    if len(df) < 4:
        return None, None, 'ข้อมูลใน sheet ไม่เพียงพอ'

    row0 = [str(v).strip() for v in df.iloc[0]]
    row1 = [str(v).strip() for v in df.iloc[1]]
    row2 = [str(v).strip() for v in df.iloc[2]]

    if 'ContactID' not in row0:
        return None, None, 'รูปแบบ Excel ไม่รองรับ — กรุณา Export จาก MyContacts ก่อน (ต้องมี ContactID ใน row 1)'

    TRUTHY = {'y', 'yes', '1', 'true', 'x', 'v'}
    def is_truthy(v): return str(v).strip().lower() in TRUTHY

    # INDIV header → DB field mapping (matches _build_fullinfo_sheet INDIV list)
    INDIV_MAP = {
        'contactid': 'id', 'name (en)': 'name_en', 'name (th)': 'name_th',
        'email 1': 'email1', 'email 2': 'email2', 'direct report to': 'direct_report',
        'team': 'team', 'sub-team': 'sub_team', 'mobile': 'phone',
        'line name': 'line_id', 'my notes': 'general_note', 'nickname': 'nickname',
        'notes': 'work_note', 'quick notes about team': 'quick_note_team',
    }
    ZONE_LABELS = {
        'projects': 'projects', 'main role': 'roles', 'entity': 'entity',
        'related systems/areas': 'systems', 'custom tags': 'tags',
    }

    # ── Scan row0 to build zone map ──────────────────────────────────────────
    indiv_cols  = {}   # db_field → col_idx
    zone_by_col = []   # per-column zone string
    cur_zone    = 'indiv'

    for ci, v in enumerate(row0):
        lv = v.lower()
        if lv in ZONE_LABELS:
            cur_zone = ZONE_LABELS[lv]
        if cur_zone == 'indiv' and lv in INDIV_MAP:
            indiv_cols[INDIV_MAP[lv]] = ci
        zone_by_col.append(cur_zone)

    # ── Projects zone: forward-fill row1 for project name, row2 for flag/sub-col ──
    proj_cols = []   # (proj_name, col_idx, is_flag, sub_col_name_or_None)
    proj_set  = set()
    last_proj = ''
    for ci, zone in enumerate(zone_by_col):
        if zone != 'projects':
            last_proj = ''
            continue
        r1 = row1[ci] if ci < len(row1) else ''
        r2 = row2[ci] if ci < len(row2) else ''
        if r1:
            last_proj = r1
        if not last_proj:
            continue
        proj_set.add(last_proj)
        if r2.lower() == last_proj.lower():
            proj_cols.append((last_proj, ci, True, None))       # flag col
        else:
            proj_cols.append((last_proj, ci, False, r2 or 'Remark'))  # sub-col

    # ── Roles zone: each column in row1 is a role name ──────────────────────
    roles_cols = [(row1[ci], ci) for ci, zone in enumerate(zone_by_col)
                  if zone == 'roles' and ci < len(row1) and row1[ci]]
    role_set = {r for r, _ in roles_cols}

    # ── Entity zone: each column in row1 is an entity type ──────────────────
    entity_cols = [(row1[ci], ci) for ci, zone in enumerate(zone_by_col)
                   if zone == 'entity' and ci < len(row1) and row1[ci]]

    # ── Systems zone ─────────────────────────────────────────────────────────
    sys_cols = [(row1[ci], ci) for ci, zone in enumerate(zone_by_col)
                if zone == 'systems' and ci < len(row1) and row1[ci]]
    sys_set = {s for s, _ in sys_cols}

    # ── Tags zone ────────────────────────────────────────────────────────────
    tag_cols = [(row1[ci], ci) for ci, zone in enumerate(zone_by_col)
                if zone == 'tags' and ci < len(row1) and row1[ci]]
    tag_set = {t for t, _ in tag_cols}

    master_in_file = {
        'projects': sorted(proj_set),
        'roles':    sorted(role_set),
        'systems':  sorted(sys_set),
        'tags':     sorted(tag_set),
    }

    def gv(vals, ci, default=''):
        if ci is None or ci >= len(vals):
            return default
        v = str(vals[ci]).strip()
        return '' if v.lower() in ('nan', '') else v

    rows = []
    for idx, (_, raw) in enumerate(df.iloc[3:].iterrows()):
        vals = [str(v).strip() for v in raw]
        if not any(v for v in vals if v and v.lower() != 'nan'):
            continue

        contact = {field: gv(vals, ci) for field, ci in indiv_cols.items()}
        cid     = contact.get('id', '')
        name_th = contact.get('name_th', '')
        if not name_th and not cid:
            continue

        # Projects
        projects_agg = {}
        for pname, ci, is_flag, sub_col in proj_cols:
            v = gv(vals, ci)
            if pname not in projects_agg:
                projects_agg[pname] = {'flag': False, 'note': '', 'subvalues': {}}
            if is_flag:
                projects_agg[pname]['flag'] = is_truthy(v)
            else:
                scn = sub_col or 'Remark'
                if scn == 'Remark':
                    projects_agg[pname]['note'] = v
                else:
                    projects_agg[pname]['subvalues'][scn] = v
        proj_list = [{'name': pn, 'note': pd_['note'], 'subvalues': pd_['subvalues']}
                     for pn, pd_ in projects_agg.items() if pd_['flag']]

        # Roles
        contact_roles = [r for r, ci in roles_cols if is_truthy(gv(vals, ci))]

        # Entity
        contact_entity = None
        for etype, ci in entity_cols:
            v = gv(vals, ci)
            if not v:
                continue
            if etype.lower() in ('vendor', 'other bank'):
                contact_entity = {'type': etype, 'value': '' if is_truthy(v) else v}
            elif is_truthy(v):
                contact_entity = {'type': etype, 'value': ''}
            if contact_entity:
                break

        # Systems
        contact_systems = [s for s, ci in sys_cols if is_truthy(gv(vals, ci))]

        # Tags
        contact_tags = [t for t, ci in tag_cols if is_truthy(gv(vals, ci))]

        rows.append({
            'row_idx':       idx,
            **contact,
            'projects':      proj_list,
            'roles':         contact_roles,
            'entity':        contact_entity,
            'systems':       contact_systems,
            'tags':          contact_tags,
        })

    return rows, master_in_file, None


@app.route('/api/import/preview', methods=['POST'])
def import_preview():
    if 'file' not in request.files:
        return jsonify({'error': 'ไม่พบไฟล์'}), 400
    f = request.files['file']
    rows, master_in_file, err = _parse_excel_for_preview(f.stream)
    if err:
        return jsonify({'error': err}), 400

    conn = db_conn()
    existing_ids   = {r['id'] for r in conn.execute("SELECT id FROM contacts").fetchall()}
    existing_email = {r['email1']: r['id'] for r in conn.execute(
        "SELECT id, email1 FROM contacts WHERE email1 != ''").fetchall()}

    db_projects = {r['name'] for r in conn.execute("SELECT name FROM master_projects").fetchall()}
    db_roles    = {r['name'] for r in conn.execute("SELECT name FROM master_roles").fetchall()}
    db_systems  = {r['name'] for r in conn.execute("SELECT name FROM master_systems").fetchall()}
    db_tags     = {r['name'] for r in conn.execute("SELECT name FROM tags").fetchall()}
    conn.close()

    new_master = {
        'projects': [p for p in master_in_file.get('projects', []) if p not in db_projects],
        'roles':    [r for r in master_in_file.get('roles',    []) if r not in db_roles],
        'systems':  [s for s in master_in_file.get('systems',  []) if s not in db_systems],
        'tags':     [t for t in master_in_file.get('tags',     []) if t not in db_tags],
    }

    INDIV_FIELDS = ['name_en','name_th','email1','email2','direct_report','team','sub_team',
                    'phone','line_id','general_note','nickname','work_note','quick_note_team']
    result = []
    for r in rows:
        cid    = r.get('id', '')
        email1 = r.get('email1', '')
        exists = (cid in existing_ids) if cid else (email1 in existing_email if email1 else False)
        existing_id = cid if cid in existing_ids else existing_email.get(email1, '')
        row_out = {
            'row_idx':     r['row_idx'],
            'contact_id':  cid or existing_id or f"row{r['row_idx']}",
            'is_new':      not exists,
            'existing_id': existing_id,
            'projects':    r.get('projects', []),
            'roles':       r.get('roles', []),
            'entity':      r.get('entity'),
            'systems':     r.get('systems', []),
            'tags':        r.get('tags', []),
        }
        for f_ in INDIV_FIELDS:
            row_out[f_] = r.get(f_, '')
        result.append(row_out)

    return jsonify({'rows': result, 'total': len(result), 'new_master': new_master})


@app.route('/api/import/execute', methods=['POST'])
def import_execute():
    """Execute import: auto-create missing master data, then upsert contacts with all zone data."""
    body            = request.get_json(force=True)
    rows            = body.get('rows', [])
    conflict_action = body.get('conflict_action', 'skip')

    conn = db_conn()
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    imported = skipped = 0

    existing_ids   = {r['id'] for r in conn.execute("SELECT id FROM contacts").fetchall()}
    existing_email = {r['email1']: r['id'] for r in conn.execute(
        "SELECT id, email1 FROM contacts WHERE email1 != ''").fetchall()}

    # ── Collect all master data referenced by selected rows ──────────────────
    all_projects, all_roles, all_systems, all_tags = set(), set(), set(), set()
    for r in rows:
        if not r.get('selected', True):
            continue
        for p in (r.get('projects') or []):
            if p.get('name'):
                all_projects.add(p['name'])
        for role in (r.get('roles') or []):
            all_roles.add(role)
        for sys_ in (r.get('systems') or []):
            all_systems.add(sys_)
        for tag in (r.get('tags') or []):
            all_tags.add(tag)

    # ── Auto-create missing master data (B+C: create + report) ──────────────
    PALETTE = ['#6366F1','#10B981','#F59E0B','#EF4444','#8B5CF6','#3B82F6','#EC4899','#14B8A6']
    created_master = {'projects': [], 'roles': [], 'systems': [], 'tags': []}

    db_projs = {r['name'] for r in conn.execute("SELECT name FROM master_projects").fetchall()}
    for pn in sorted(all_projects):
        if pn not in db_projs:
            c = PALETTE[len(created_master['projects']) % len(PALETTE)]
            conn.execute("INSERT OR IGNORE INTO master_projects(name,color,short_name) VALUES(?,?,?)",
                         (pn, c, pn[:4]))
            created_master['projects'].append(pn)

    db_roles = {r['name'] for r in conn.execute("SELECT name FROM master_roles").fetchall()}
    for rn in sorted(all_roles):
        if rn not in db_roles:
            c = PALETTE[(len(created_master['roles']) + 3) % len(PALETTE)]
            conn.execute("INSERT OR IGNORE INTO master_roles(name,color) VALUES(?,?)", (rn, c))
            created_master['roles'].append(rn)

    db_sys = {r['name'] for r in conn.execute("SELECT name FROM master_systems").fetchall()}
    for sn in sorted(all_systems):
        if sn not in db_sys:
            conn.execute("INSERT OR IGNORE INTO master_systems(name) VALUES(?)", (sn,))
            created_master['systems'].append(sn)

    db_tags_set = {r['name'] for r in conn.execute("SELECT name FROM tags").fetchall()}
    for tn in sorted(all_tags):
        if tn not in db_tags_set:
            conn.execute("INSERT OR IGNORE INTO tags(name,color) VALUES(?,?)", (tn, '#64748B'))
            created_master['tags'].append(tn)

    conn.commit()
    tag_id_map = {r['name']: r['id'] for r in conn.execute("SELECT id, name FROM tags").fetchall()}

    # ── Upsert contacts ───────────────────────────────────────────────────────
    FIELDS = ['name_en','email1','email2','team','sub_team','org_role','direct_report',
              'phone','line_id','nickname','note_short','general_note','work_note','quick_note_team']

    for r in rows:
        if not r.get('selected', True):
            skipped += 1
            continue
        cid     = r.get('contact_id', '')
        is_new  = r.get('is_new', True)
        name_th = (r.get('name_th') or '').strip()
        if not name_th:
            skipped += 1
            continue
        if not is_new and conflict_action == 'skip':
            skipped += 1
            continue

        if not cid or (not is_new and cid not in existing_ids):
            cid = (r.get('existing_id') or existing_email.get(r.get('email1',''), '') or
                   _next_id(conn))

        vals_dict = {f: (r.get(f) or '') for f in FIELDS}

        if not is_new and cid in existing_ids:
            sets = ', '.join(f"{f}=?" for f in FIELDS) + ', name_th=?, updated_at=?'
            conn.execute(f"UPDATE contacts SET {sets} WHERE id=?",
                         [vals_dict[f] for f in FIELDS] + [name_th, now, cid])
        else:
            conn.execute("""INSERT OR IGNORE INTO contacts
                (id,name_th,name_en,email1,email2,team,sub_team,org_role,direct_report,
                 phone,line_id,nickname,note_short,general_note,work_note,quick_note_team,
                 created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cid, name_th, vals_dict['name_en'], vals_dict['email1'], vals_dict['email2'],
                 vals_dict['team'], vals_dict['sub_team'], vals_dict['org_role'],
                 vals_dict['direct_report'], vals_dict['phone'], vals_dict['line_id'],
                 vals_dict['nickname'], vals_dict['note_short'], vals_dict['general_note'],
                 vals_dict['work_note'], vals_dict['quick_note_team'], now, now))

        # Projects
        conn.execute("DELETE FROM contact_projects WHERE contact_id=?", (cid,))
        conn.execute("DELETE FROM contact_project_subvalues WHERE contact_id=?", (cid,))
        for i, p in enumerate(r.get('projects') or []):
            pname = p.get('name', '')
            if not pname:
                continue
            conn.execute(
                "INSERT INTO contact_projects(contact_id,project_name,role,note,sort_order) VALUES(?,?,?,?,?)",
                (cid, pname, p.get('role', 'Supporter'), p.get('note', ''), i))
            for scn, scv in (p.get('subvalues') or {}).items():
                if scv and str(scv).strip():
                    conn.execute(
                        "INSERT OR REPLACE INTO contact_project_subvalues VALUES(?,?,?,?)",
                        (cid, pname, scn, str(scv).strip()))

        # Roles
        conn.execute("DELETE FROM contact_roles WHERE contact_id=?", (cid,))
        for role in (r.get('roles') or []):
            conn.execute("INSERT OR IGNORE INTO contact_roles VALUES(?,?)", (cid, role))

        # Entity
        conn.execute("DELETE FROM contact_entity WHERE contact_id=?", (cid,))
        ent = r.get('entity')
        if ent and ent.get('type'):
            conn.execute("INSERT INTO contact_entity VALUES(?,?,?)",
                         (cid, ent['type'], ent.get('value') or ''))

        # Systems
        conn.execute("DELETE FROM contact_systems WHERE contact_id=?", (cid,))
        for sys_ in (r.get('systems') or []):
            conn.execute("INSERT OR IGNORE INTO contact_systems VALUES(?,?)", (cid, sys_))

        # Tags
        conn.execute("DELETE FROM contact_tags WHERE contact_id=?", (cid,))
        for tname in (r.get('tags') or []):
            tid = tag_id_map.get(tname)
            if tid:
                conn.execute("INSERT OR IGNORE INTO contact_tags VALUES(?,?)", (cid, tid))

        imported += 1

    conn.commit()
    conn.close()
    return jsonify({'imported': imported, 'skipped': skipped, 'created_master': created_master})


# ─── Meeting Room ──────────────────────────────────────────────────────────────

@app.route('/meeting')
def meeting_room():
    return render_template('meeting_room.html')


def _parse_meeting_datetime(dt_str):
    """Parse 'Thu 30 Apr 26 10:00 - 11:00' → '2026.04.30(Thu) 10:00 - 11:00'"""
    import re
    if not dt_str:
        return ''
    MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
              'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    m = re.match(
        r'(\w{3})[,\s]+(\d{1,2})\s+(\w{3,})\s+(\d{2,4})\s+(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})',
        dt_str.strip(), re.I)
    if m:
        day_name, day, mon_str, year, t_start, t_end = m.groups()
        mon_num   = MONTHS.get(mon_str.lower()[:3], 1)
        year_full = int(year) + (2000 if len(year) == 2 else 0)
        day_abbr  = day_name[:3].capitalize()
        return f"{year_full}.{mon_num:02d}.{int(day):02d}({day_abbr}) {t_start} - {t_end}"
    return dt_str


def _parse_attendee_line(line):
    """Classify one attendee line → dict with type/name/email/entity."""
    import re
    line = line.strip()
    if not line:
        return None

    # Pattern 1: "Firstname Lastname (Organization)"
    m = re.match(r'^(.+?)\s+\((.+?)\)$', line)
    if m:
        name_part = m.group(1).strip()
        parts     = name_part.split()
        email     = '.'.join(parts) + '@krungsri.com'
        return {'type': 'person', 'name': name_part, 'email': email,
                'entity': 'KS', 'email_derived': True}

    # Pattern 2: direct email
    if '@' in line:
        email = line.strip()
        local = email.split('@')[0]
        name  = ' '.join(p.capitalize() for p in local.split('.') if p)
        return {'type': 'person', 'name': name, 'email': email,
                'entity': 'KS', 'email_derived': False}

    # Pattern 3: group mail (contains _ or no spaces)
    if '_' in line or ' ' not in line:
        return {'type': 'group', 'name': line,
                'email': line + '@krungsri.com', 'entity': 'KS', 'email_derived': True}

    # Fallback: treat as plain name
    parts = line.split()
    return {'type': 'person', 'name': line,
            'email': '.'.join(parts) + '@krungsri.com',
            'entity': 'KS', 'email_derived': True}


def _parse_invite_text(text):
    subject, dt_raw, attendee_lines = '', '', []
    in_att = False
    for raw in text.strip().splitlines():
        line = raw.strip()
        low  = line.lower()
        if low.startswith('subject:'):
            subject = line[8:].strip()
        elif low.startswith('datetime:'):
            dt_raw = line[9:].strip()
        elif low.startswith('attendees:') or low.startswith('required:') or low.startswith('optional:'):
            in_att = True
        elif in_att and line:
            attendee_lines.append(line)

    date_label = _parse_meeting_datetime(dt_raw)
    atm_prefix = f"{date_label} >> {subject}".strip(' >')
    return {
        'subject':      subject,
        'datetime_raw': dt_raw,
        'date_label':   date_label,
        'atm_prefix':   atm_prefix,
        'attendees':    [a for a in (_parse_attendee_line(l) for l in attendee_lines) if a],
    }


@app.route('/api/meeting/parse', methods=['POST'])
def meeting_parse():
    body    = request.get_json(force=True)
    parsed  = _parse_invite_text(body.get('text', ''))
    conn    = db_conn()
    found, new_list, groups = [], [], []

    for att in parsed['attendees']:
        if att['type'] == 'group':
            groups.append(att)
            continue

        email_lo = att['email'].lower()
        name_lo  = att['name'].lower()
        row = conn.execute(
            "SELECT * FROM contacts WHERE LOWER(email1)=? OR LOWER(email2)=? OR LOWER(name_en)=?",
            (email_lo, email_lo, name_lo)).fetchone()

        if row:
            cd = _contact_dict(row, conn, full=False)
            cd['input_name']    = att['name']
            cd['derived_email'] = att['email']
            cd['email_derived'] = att['email_derived']
            found.append(cd)
        else:
            new_list.append(att)

    conn.close()
    return jsonify({
        'meeting': {
            'subject':      parsed['subject'],
            'datetime_raw': parsed['datetime_raw'],
            'date_label':   parsed['date_label'],
            'atm_prefix':   parsed['atm_prefix'],
        },
        'found':  found,
        'new':    new_list,
        'groups': groups,
    })


@app.route('/api/meeting/add', methods=['POST'])
def meeting_add():
    """Batch-create new contacts discovered in a meeting."""
    body     = request.get_json(force=True)
    contacts = body.get('contacts', [])
    atm_line = body.get('atm_line', '').strip()
    conn     = db_conn()
    created, errors = [], []
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for cd in contacts:
        try:
            cid  = _next_id(conn)
            name = (cd.get('name') or '').strip()
            if not name:
                errors.append({'name': '(empty)', 'error': 'ไม่มีชื่อ'})
                continue

            conn.execute("""INSERT INTO contacts
                (id,name_th,name_en,nickname,team,sub_team,org_role,
                 email1,note_short,general_note,work_note,quick_note_team,
                 associated_to_meeting,created_at,updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cid,
                 cd.get('name_th') or name,
                 cd.get('name_en') or name,
                 cd.get('nickname',''),
                 cd.get('team',''),
                 cd.get('sub_team',''),
                 cd.get('org_role',''),
                 cd.get('email',''),
                 cd.get('note_short',''),
                 cd.get('general_note',''),
                 cd.get('work_note',''),
                 cd.get('quick_note_team',''),
                 atm_line,
                 now, now))

            etype = cd.get('entity','KS')
            if etype:
                conn.execute("INSERT OR REPLACE INTO contact_entity VALUES (?,?,?)",
                             (cid, etype, ''))
            for s in (cd.get('systems') or []):
                conn.execute("INSERT OR IGNORE INTO contact_systems VALUES (?,?)", (cid, s))
            for r in (cd.get('roles') or []):
                conn.execute("INSERT OR IGNORE INTO contact_roles VALUES (?,?)", (cid, r))
            for i, p in enumerate(cd.get('projects') or []):
                conn.execute(
                    "INSERT INTO contact_projects (contact_id,project_name,role,note,sort_order) VALUES (?,?,?,?,?)",
                    (cid, p['project_name'], p.get('role','Supporter'), p.get('note',''), i))
            conn.commit()
            created.append(cid)
        except Exception as e:
            errors.append({'name': cd.get('name','?'), 'error': str(e)})

    conn.close()
    return jsonify({'created': created, 'errors': errors})


@app.route('/api/contacts/<cid>/atm', methods=['POST'])
def append_atm(cid):
    """Append one line to a contact's associated_to_meeting field."""
    body = request.get_json(force=True)
    line = (body.get('line') or '').strip()
    if not line:
        return jsonify({'error': 'missing line'}), 400
    conn = db_conn()
    row  = conn.execute("SELECT associated_to_meeting FROM contacts WHERE id=?", (cid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    current = (row['associated_to_meeting'] or '').strip()
    new_val = (current + '\n' + line).strip() if current else line
    conn.execute("UPDATE contacts SET associated_to_meeting=?, updated_at=? WHERE id=?",
                 (new_val, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), cid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'associated_to_meeting': new_val})


@app.route('/api/master/teams', methods=['POST'])
def create_team():
    body = request.get_json(force=True)
    name = (body.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    conn = db_conn()
    conn.execute("INSERT OR IGNORE INTO master_teams (name, color) VALUES (?,?)",
                 (name, body.get('color','#64748B')))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/db/clear', methods=['POST'])
def clear_db():
    body = request.get_json(force=True)
    if not body.get('confirm'):
        return jsonify({'error': 'confirm required'}), 400
    conn = db_conn()
    n = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
    for tbl in ['contact_project_subvalues','contact_projects','contact_tags',
                'contact_roles','contact_entity','contact_systems','notes','contacts']:
        conn.execute(f"DELETE FROM {tbl}")
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'deleted': n})


# ─── Run ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    watcher.start()
    print("MyContacts running at http://localhost:5000")
    app.run(debug=True, port=5000, use_reloader=False)
